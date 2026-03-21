import telebot
from telebot import types, apihelper
from flask import Flask
from threading import Thread, Lock
import psycopg2
from psycopg2 import pool as psycopg2_pool
from datetime import datetime, timedelta
from collections import deque
import time
import os
import json
import logging
import re
import io
import math
import functools

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ── ENV ───────────────────────────────────────────────────────
TOKEN = os.environ.get("BOT_TOKEN")
if not TOKEN:
    raise RuntimeError("BOT_TOKEN environment variable is not set!")
DATABASE_URL = os.environ.get("DATABASE_URL") or ""

def _parse_admin_ids():
    env_val = os.environ.get("ADMIN_IDS", "")
    if env_val:
        try:
            return set(int(x.strip()) for x in env_val.split(",") if x.strip())
        except Exception:
            pass
    return {5880534778, 5541976681}

ADMIN_IDS = _parse_admin_ids()
def is_admin(uid): return uid in ADMIN_IDS

bot = telebot.TeleBot(TOKEN, parse_mode="HTML")

# ── DB POOL ───────────────────────────────────────────────────
_db_pool = None
_db_pool_lock = Lock()

def get_pool():
    global _db_pool
    if _db_pool is None:
        with _db_pool_lock:
            if _db_pool is None:
                if not DATABASE_URL:
                    raise RuntimeError("DATABASE_URL орнатылмаған!")
                _db_pool = psycopg2_pool.ThreadedConnectionPool(
                    minconn=2, maxconn=15, dsn=DATABASE_URL, connect_timeout=10)
    return _db_pool

def get_db():
    p = get_pool()
    conn = p.getconn()
    return conn, conn.cursor()

def release_db(conn):
    try:
        get_pool().putconn(conn)
    except Exception as e:
        logger.warning(f"release_db: {e}")

# FIX: Güvenli context manager для DB
from contextlib import contextmanager

@contextmanager
def db_cursor():
    """
    Автоматически освобождает connection и cursor.
    Использование:
        with db_cursor() as (conn, cur):
            cur.execute(...)
            conn.commit()
    """
    conn, cursor = get_db()
    try:
        yield conn, cursor
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        release_db(conn)

# ── INIT DB ───────────────────────────────────────────────────
def init_db():
    tables = [
        """CREATE TABLE IF NOT EXISTS students (
            id BIGINT PRIMARY KEY, username TEXT, last_active TIMESTAMP,
            full_name TEXT, birth_date TEXT, phone TEXT, hemis TEXT, started INTEGER DEFAULT 0)""",
        """CREATE TABLE IF NOT EXISTS user_news (
            id SERIAL PRIMARY KEY, content TEXT, author_id BIGINT,
            author_username TEXT, date TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS materials (
            id SERIAL PRIMARY KEY, file_id TEXT, file_type TEXT,
            uploader_id BIGINT, uploader_username TEXT, date TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS gallery (
            id SERIAL PRIMARY KEY, file_id TEXT, file_type TEXT,
            uploader_id BIGINT, uploader_username TEXT, date TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS schedule (
            id SERIAL PRIMARY KEY, day TEXT, subject TEXT, time TEXT)""",
        """CREATE TABLE IF NOT EXISTS suggestions (
            id SERIAL PRIMARY KEY, content TEXT, user_id BIGINT, date TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY, date TEXT, para INTEGER, subject TEXT,
            student_id BIGINT, student_name TEXT, status TEXT, marked_at TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS contacts (
            id SERIAL PRIMARY KEY, type TEXT, name TEXT, phone TEXT)""",
        """CREATE TABLE IF NOT EXISTS contracts (
            id SERIAL PRIMARY KEY, student_id BIGINT UNIQUE, total_amount REAL, note TEXT)""",
        """CREATE TABLE IF NOT EXISTS contract_payments (
            id SERIAL PRIMARY KEY, student_id BIGINT, amount REAL, date TEXT, note TEXT)""",
        """CREATE TABLE IF NOT EXISTS test_variants (
            id SERIAL PRIMARY KEY, subject TEXT, file_id TEXT, file_type TEXT,
            file_name TEXT, uploader_id BIGINT, date TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS user_states (
            user_id BIGINT PRIMARY KEY, state TEXT, updated_at TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS attendance_sessions (
            admin_id BIGINT PRIMARY KEY, session_data TEXT, updated_at TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS blocked_users (
            user_id BIGINT PRIMARY KEY, reason TEXT, blocked_at TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS ai_history (
            user_id BIGINT PRIMARY KEY, history TEXT, updated_at TIMESTAMP DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS sent_reminders (
            key TEXT PRIMARY KEY, sent_at TIMESTAMP DEFAULT NOW())""",
    ]
    migrations = [
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS full_name TEXT",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS birth_date TEXT",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS phone TEXT",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS hemis TEXT",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS started INTEGER DEFAULT 0",
    ]
    with db_cursor() as (conn, cursor):
        for sql in tables:
            cursor.execute(sql)
        for sql in migrations:
            try:
                cursor.execute(sql)
            except Exception:
                conn.rollback()
        conn.commit()

init_db()

UZ_OFFSET = timedelta(hours=5)
def now_uz(): return datetime.utcnow() + UZ_OFFSET

# ── FLASK ─────────────────────────────────────────────────────
app = Flask(__name__)

@app.route("/")
def home(): return "Bot is alive", 200

@app.route("/health")
def health(): return {"status": "ok", "time": str(now_uz())}, 200

@app.route("/ping")
def ping(): return "pong", 200

# ── RATE LIMIT ────────────────────────────────────────────────
_rate_limit = {}
_rate_limit_lock = Lock()
RATE_LIMIT_MAX = 50
RATE_LIMIT_WINDOW = 30

def is_rate_limited(uid):
    if is_admin(uid): return False
    now = time.time()
    with _rate_limit_lock:
        h = [t for t in _rate_limit.get(uid, []) if now - t < RATE_LIMIT_WINDOW]
        h.append(now)
        _rate_limit[uid] = h
        return len(h) > RATE_LIMIT_MAX

def clean_rate_limit():
    now = time.time()
    with _rate_limit_lock:
        for uid in list(_rate_limit):
            if all(now - t > RATE_LIMIT_WINDOW * 2 for t in _rate_limit[uid]):
                del _rate_limit[uid]

# ── ACCESS HELPERS ────────────────────────────────────────────
# FIX: is_blocked/is_authorized әр хабарламада DB сұранысы жасайды.
# Кіші бот үшін blocked_users кэші қосылды (өзгерісте автоматты жаңарады).
_blocked_cache: set = set()   # блокланған user_id жиыны
_blocked_cache_lock = Lock()
_blocked_cache_loaded = False

def _load_blocked_cache():
    global _blocked_cache_loaded
    with _blocked_cache_lock:
        if _blocked_cache_loaded:
            return
    try:
        with db_cursor() as (_, cursor):
            cursor.execute("SELECT user_id FROM blocked_users")
            ids = {r[0] for r in cursor.fetchall()}
        with _blocked_cache_lock:
            _blocked_cache.clear()
            _blocked_cache.update(ids)
            _blocked_cache_loaded = True
    except Exception as e:
        logger.warning(f"_load_blocked_cache: {e}")

def is_blocked(uid):
    _load_blocked_cache()
    with _blocked_cache_lock:
        return uid in _blocked_cache

def _add_to_blocked_cache(uid):
    with _blocked_cache_lock:
        _blocked_cache.add(uid)

def _remove_from_blocked_cache(uid):
    with _blocked_cache_lock:
        _blocked_cache.discard(uid)

def is_authorized(uid):
    if is_admin(uid): return True
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id FROM students WHERE id=%s", (uid,))
        return cursor.fetchone() is not None

def _update_last_active(uid):
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("UPDATE students SET last_active=%s WHERE id=%s", (now_uz(), uid))
            conn.commit()
    except Exception:
        pass

def check_access(func):
    @functools.wraps(func)
    def wrapper(message):
        uid = message.from_user.id
        if is_blocked(uid):
            try: bot.send_message(uid, "⛔ Сиз блокландыңыз. Admin-ге хабарласыңыз.")
            except Exception: pass
            return
        if message.text != "/start" and not is_admin(uid) and not is_authorized(uid):
            try: bot.send_message(uid, "⛔ <b>Кириуге рұхсат жоқ!</b>\nАдминге хабарласыңыз.")
            except Exception: pass
            return
        if is_rate_limited(uid):
            try: bot.send_message(uid, "⏳ Дым тез! Бираздан кейин қайталаңыз.")
            except Exception: pass
            return
        if not is_admin(uid):
            _update_last_active(uid)
        return func(message)
    return wrapper

def check_access_cb(func):
    @functools.wraps(func)
    def wrapper(call):
        uid = call.from_user.id
        if is_blocked(uid):
            bot.answer_callback_query(call.id, "⛔ Сиз блокландыңыз.")
            return
        if not is_admin(uid) and not is_authorized(uid):
            bot.answer_callback_query(call.id, "⛔ Рұхсат жоқ!")
            return
        if not is_admin(uid):
            _update_last_active(uid)
        return func(call)
    return wrapper

def user_step_check(message):
    uid = message.from_user.id
    if is_blocked(uid):
        try: bot.send_message(uid, "⛔ Сиз блокландыңыз.")
        except: pass
        return False
    if not is_authorized(uid) and not is_admin(uid):
        try: bot.send_message(uid, "⛔ Рұхсат жоқ!")
        except: pass
        return False
    return True

# ── CONSTANTS ─────────────────────────────────────────────────
DAYS_RU = ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]
DAYS_EN_TO_RU = {
    "Monday":"Понедельник","Tuesday":"Вторник","Wednesday":"Среда",
    "Thursday":"Четверг","Friday":"Пятница","Saturday":"Суббота","Sunday":"Воскресенье"}
MONTHS_RU = {
    1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
    7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
WEEKDAYS_RU = {0:"Понедельник",1:"Вторник",2:"Среда",3:"Четверг",
               4:"Пятница",5:"Суббота",6:"Воскресенье"}

# FIX: SQL injection-дан қорғау үшін рұхсат етілген кестелер
ALLOWED_DELETE_TABLES = {"materials", "gallery", "user_news"}

def clean_hemis(val):
    if val is None: return ""
    if isinstance(val, float) and math.isnan(val): return ""
    s = str(val).strip()
    if s in ("None", "nan", ""): return ""
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): return s[:-2]
    return s

def get_birthday_info(birth_date_str):
    try:
        if not birth_date_str: return None, None
        s = str(birth_date_str).strip()
        if not s or s in ("None", "nan", ""): return None, None
        bd = datetime.strptime(s[:10], "%Y-%m-%d")
        today = now_uz().date()
        try:
            this_year_bd = bd.replace(year=today.year).date()
        except ValueError:
            this_year_bd = bd.replace(year=today.year, day=28).date()
        if this_year_bd == today: return 0, today
        if this_year_bd < today:
            try:
                this_year_bd = bd.replace(year=today.year + 1).date()
            except ValueError:
                this_year_bd = bd.replace(year=today.year + 1, day=28).date()
        return (this_year_bd - today).days, this_year_bd
    except Exception as e:
        logger.warning(f"get_birthday_info({birth_date_str}): {e}")
        return None, None

# ── STATE ─────────────────────────────────────────────────────
# FIX: handler lambda-лар әр хабарламада get_user_state() шақырады
# → DB bottleneck болмау үшін in-memory кэш қосылды.
# DB — бот рестарт болса қалпына келтіру үшін сақтайды.
_state_cache: dict = {}       # {uid: state_str or None}
_state_cache_lock = Lock()

def set_user_state(uid, state):
    with _state_cache_lock:
        _state_cache[uid] = state
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO user_states(user_id,state,updated_at) VALUES(%s,%s,%s) "
                "ON CONFLICT(user_id) DO UPDATE SET state=excluded.state,updated_at=excluded.updated_at",
                (uid, state, now_uz()))
            conn.commit()
    except Exception as e:
        logger.warning(f"set_user_state DB: {e}")

def get_user_state(uid):
    with _state_cache_lock:
        if uid in _state_cache:
            return _state_cache[uid]
    # Кэште жоқ болса — DB-ден оқу (рестарттан кейін)
    try:
        with db_cursor() as (_, cursor):
            cursor.execute("SELECT state FROM user_states WHERE user_id=%s", (uid,))
            row = cursor.fetchone()
        val = row[0] if row else None
        with _state_cache_lock:
            _state_cache[uid] = val
        return val
    except Exception as e:
        logger.warning(f"get_user_state DB: {e}")
        return None

def clear_user_state(uid):
    with _state_cache_lock:
        _state_cache[uid] = None
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("DELETE FROM user_states WHERE user_id=%s", (uid,))
            conn.commit()
    except Exception as e:
        logger.warning(f"clear_user_state DB: {e}")

# ── ATTENDANCE SESSION ────────────────────────────────────────
def save_attendance_session(admin_id, session):
    with db_cursor() as (conn, cursor):
        cursor.execute(
            "INSERT INTO attendance_sessions(admin_id,session_data,updated_at) VALUES(%s,%s,%s) "
            "ON CONFLICT(admin_id) DO UPDATE SET session_data=excluded.session_data,updated_at=excluded.updated_at",
            (admin_id, json.dumps(session, ensure_ascii=False), now_uz()))
        conn.commit()

def load_attendance_session(admin_id):
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT session_data FROM attendance_sessions WHERE admin_id=%s", (admin_id,))
        row = cursor.fetchone()
    if not row: return None
    s = json.loads(row[0])
    if "results" in s:
        s["results"] = {int(k): v for k, v in s["results"].items()}
    return s

def delete_attendance_session(admin_id):
    with db_cursor() as (conn, cursor):
        cursor.execute("DELETE FROM attendance_sessions WHERE admin_id=%s", (admin_id,))
        conn.commit()

def cleanup_old_sessions():
    with db_cursor() as (conn, cursor):
        cursor.execute("DELETE FROM attendance_sessions WHERE updated_at < %s",
                       (now_uz() - timedelta(hours=2),))
        conn.commit()

# ── DATE HELPERS ──────────────────────────────────────────────
def date_to_ru(ds):
    try:
        dt = datetime.strptime(str(ds)[:10], "%Y-%m-%d")
        return f"{WEEKDAYS_RU[dt.weekday()]}, {dt.day} {MONTHS_RU[dt.month]}"
    except:
        return str(ds)

def get_online_status(la):
    try:
        last = la if isinstance(la, datetime) else datetime.strptime(str(la)[:19], "%Y-%m-%d %H:%M:%S")
        d = max((now_uz() - last).total_seconds(), 0)
        if d < 900: return "🟢 Онлайн"
        elif d < 3600: return f"🟡 {int(d//60)} мин бұрын"
        elif d < 86400: return f"🔴 {int(d//3600)} сағ бұрын"
        else: return f"🔴 {int(d//86400)} күн бұрын"
    except:
        return "⚪ Белгисиз"

def _is_online(la, now_t):
    try:
        last = la if isinstance(la, datetime) else datetime.strptime(str(la)[:19], "%Y-%m-%d %H:%M:%S")
        return (now_t - last).total_seconds() < 900
    except:
        return False

# ── MESSAGE HELPERS ───────────────────────────────────────────
def send_long_message(chat_id, text, reply_markup=None, chunk_size=3800):
    if len(text) <= chunk_size:
        bot.send_message(chat_id, text, reply_markup=reply_markup)
        return
    lines = text.split("\n")
    parts = []
    cur = ""
    for line in lines:
        if len(cur) + len(line) + 1 > chunk_size:
            if cur: parts.append(cur)
            cur = line
        else:
            cur = cur + "\n" + line if cur else line
    if cur: parts.append(cur)
    for i, p in enumerate(parts):
        bot.send_message(chat_id, p, reply_markup=reply_markup if i == len(parts) - 1 else None)

def send_to_students(text=None, file_id=None, file_type=None, exclude_id=None):
    with db_cursor() as (_, cursor):
        if exclude_id:
            cursor.execute("SELECT id FROM students WHERE started=1 AND id!=%s", (exclude_id,))
        else:
            cursor.execute("SELECT id FROM students WHERE started=1")
        rows = cursor.fetchall()

    def _do():
        deactivated = []
        for (sid,) in rows:
            try:
                if file_id and file_type == "photo":
                    bot.send_photo(sid, file_id, caption=text)
                elif file_id and file_type == "document":
                    bot.send_document(sid, file_id, caption=text)
                elif file_id and file_type == "video":
                    bot.send_video(sid, file_id, caption=text)
                elif text:
                    bot.send_message(sid, text)
                time.sleep(0.05)
            except apihelper.ApiTelegramException as e:
                err = str(e).lower()
                # FIX: 'not found' тым кең — басқа қателерді жұтып алады.
                # Нақты Telegram error code-тарын тексеру:
                if any(x in err for x in [
                    "bot was blocked by the user",
                    "user is deactivated",
                    "chat not found",
                    "have no rights",
                    "forbidden",
                ]) or "403" in err:
                    deactivated.append(sid)
                    logger.info(f"send_to_students: {sid} бот-ты блоклаған немесе белсенді емес")
                else:
                    logger.warning(f"send_to_students({sid}): {e}")
            except Exception as e:
                logger.warning(f"send_to_students({sid}): {e}")

        if deactivated:
            try:
                with db_cursor() as (conn2, cur2):
                    for sid in deactivated:
                        cur2.execute("UPDATE students SET started=0 WHERE id=%s", (sid,))
                    conn2.commit()
            except Exception as e:
                logger.warning(f"deactivated update: {e}")

    Thread(target=_do, daemon=True).start()

def send_to_all_students(text=None, exclude_id=None):
    with db_cursor() as (_, cursor):
        if exclude_id:
            cursor.execute("SELECT id FROM students WHERE id!=%s", (exclude_id,))
        else:
            cursor.execute("SELECT id FROM students")
        rows = cursor.fetchall()

    def _do():
        for (sid,) in rows:
            try:
                if text: bot.send_message(sid, text)
                time.sleep(0.05)
            except apihelper.ApiTelegramException as e:
                err = str(e).lower()
                if any(x in err for x in ["blocked", "403", "chat not found", "user is deactivated"]):
                    logger.info(f"send_to_all: {sid} белсенди емес")
                else:
                    logger.warning(f"send_to_all({sid}): {e}")
            except Exception as e:
                logger.warning(f"send_to_all({sid}): {e}")

    Thread(target=_do, daemon=True).start()

_processed_messages = deque(maxlen=500)
_processed_lock = Lock()

def is_already_processed(mid):
    with _processed_lock:
        if mid in _processed_messages:
            return True
        _processed_messages.append(mid)
        return False

_last_saved = {}
_last_saved_lock = Lock()

def send_saved_once(chat_id, uid):
    now = time.time()
    with _last_saved_lock:
        send = now - _last_saved.get(uid, 0) > 30
        if send: _last_saved[uid] = now
    if send: bot.send_message(chat_id, "✅ <b>Сақланды!</b>")

# ── EXCEL ATTENDANCE ──────────────────────────────────────────
def generate_attendance_excel(students, results, date_str, para, subject):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl орнатылмаған!")
        return None
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Барлау"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="2E75B6")
        ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
        la = Alignment(horizontal="left", vertical="center", wrap_text=True)
        gf = PatternFill("solid", fgColor="C6EFCE")
        rf = PatternFill("solid", fgColor="FFC7CE")
        tb = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
        headers = ["№", "ФИО", "Күн", "Пара", "Пән", "Барлау"]
        widths = [5, 38, 14, 8, 22, 12]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ca
            cell.border = tb
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 20
        rn = 2
        for i, item in enumerate(students, 1):
            if not isinstance(item, (list, tuple)) or len(item) < 2:
                continue
            first, second = item[0], item[1]
            if isinstance(first, int):
                name = str(second) if second else "—"
                status = results.get(first, "absent") if results else "absent"
            else:
                name = str(first) if first else "—"
                status = str(second) if second else "absent"
            st_text = "✅ Бар" if status == "present" else "❌ Жоқ"
            rfill = gf if status == "present" else rf
            for ci, val in enumerate([i, name, date_str, para, subject, st_text], 1):
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.fill = rfill
                cell.border = tb
                cell.alignment = la if ci == 2 else ca
            rn += 1
        ws.auto_filter.ref = f"A1:F{rn-1}"
        safe_date = date_str.replace("-", "")
        safe_subj = re.sub(r'[^\w]', '_', subject)[:20]
        path = f"/tmp/attendance_{safe_date}_para{para}_{safe_subj}.xlsx"
        wb.save(path)
        return path
    except Exception as e:
        logger.error(f"generate_attendance_excel: {e}", exc_info=True)
        return None

def send_excel_file(chat_id, path, caption=""):
    try:
        with open(path, "rb") as f:
            data = f.read()
        try:
            os.remove(path)
        except:
            pass
        fo = io.BytesIO(data)
        fo.name = os.path.basename(path)
        bot.send_document(chat_id, fo, caption=caption)
        return True
    except Exception as e:
        logger.error(f"send_excel_file: {e}")
        return False

# ── EXCEL DOWNLOAD ────────────────────────────────────────────
def _excel_download_impl(message):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        bot.send_message(message.chat.id,
            "❌ openpyxl орнатылмаған!\n<code>pip install openpyxl</code>",
            reply_markup=excel_submenu())
        return
    try:
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT id,full_name,birth_date,phone,hemis,username FROM students ORDER BY full_name")
            rows = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Студентлер"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="2E75B6")
        ca = Alignment(horizontal="center", vertical="center")
        tb = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
        alt = PatternFill("solid", fgColor="EBF3FB")
        headers = ["№", "ФИО", "Тууылған күни", "Телефон", "HEMIS", "Telegram", "TelegramID"]
        widths = [5, 38, 16, 20, 18, 22, 16]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ca
            cell.border = tb
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 22

        for i, row in enumerate(rows, 1):
            tg_id, full_name, birth_date, phone, hemis, username = row
            hemis_val = clean_hemis(hemis)
            uname = f"@{username}" if username else ""
            rfill = alt if i % 2 == 0 else None
            for ci, val in enumerate([i, full_name or "", birth_date or "", phone or "", hemis_val, uname, tg_id], 1):
                cell = ws.cell(row=i + 1, column=ci, value=val)
                cell.border = tb
                cell.alignment = ca
                if rfill: cell.fill = rfill

        for r in range(2, len(rows) + 2):
            ws.cell(row=r, column=7).number_format = '0'
            cell_bd = ws.cell(row=r, column=3)
            bd_val = cell_bd.value
            if bd_val:
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
                    try:
                        cell_bd.value = datetime.strptime(str(bd_val), fmt)
                        cell_bd.number_format = 'DD.MM.YYYY'
                        break
                    except:
                        pass

        ws.auto_filter.ref = f"A1:G{len(rows)+1}"
        ws.freeze_panes = "A2"
        path = "/tmp/students_export.xlsx"
        wb.save(path)
        send_excel_file(message.chat.id, path, caption=(
            f"📥 <b>Студентлер дизими</b> — {len(rows)} студент\n\n"
            "✏️ <b>Толтырыу қателиги:</b>\n"
            "   B — ФИО\n   C — Тууылған күни (2000-01-15)\n"
            "   D — Телефон\n   E — HEMIS ID\n\n"
            "⚠️ <b>G қатарын (TelegramID) өзгертпеңиз!</b>\n"
            "📤 Толтырып болғаннан кейин <b>Excel импорт</b> арқалы жүклеңиз."))
        bot.send_message(message.chat.id, "✅ Жиберилди!", reply_markup=excel_submenu())
    except Exception as e:
        logger.error(f"excel_download: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Excel жасауда қате: {e}", reply_markup=excel_submenu())

# ── EXCEL IMPORT ──────────────────────────────────────────────
def _excel_import_impl(message):
    if not message.document:
        msg = bot.send_message(message.chat.id, "⚠️ .xlsx файл жибериңиз!", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_excel_import)
        return
    fname = message.document.file_name or ""
    if not fname.lower().endswith(".xlsx"):
        msg = bot.send_message(message.chat.id, "⚠️ Тек .xlsx форматы қабылланады!", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_excel_import)
        return
    try:
        import openpyxl
    except ImportError:
        bot.send_message(message.chat.id,
            "❌ openpyxl орнатылмаған!\n<code>pip install openpyxl</code>",
            reply_markup=excel_submenu())
        return

    path = "/tmp/import_students.xlsx"
    try:
        fi = bot.get_file(message.document.file_id)
        with open(path, "wb") as f:
            f.write(bot.download_file(fi.file_path))
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Файлды жүклеу мүмкин болмады: {e}", reply_markup=excel_submenu())
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Excel файлды оқыу мүмкин болмады: {e}", reply_markup=excel_submenu())
        try: os.remove(path)
        except: pass
        return

    def clean_cell(val):
        if val is None: return ""
        if isinstance(val, float) and math.isnan(val): return ""
        s = str(val).strip()
        if s in ("None", "nan", ""): return ""
        if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): return s[:-2]
        return s

    def parse_birth_date(val):
        if val is None: return ""
        if hasattr(val, 'strftime'):
            try: return val.strftime("%Y-%m-%d")
            except: return ""
        s = str(val).strip()
        if not s or s in ("None", "nan", ""): return ""
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y", "%d-%m-%Y", "%Y%m%d"):
            try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except: pass
        return s

    def parse_tg_id(val):
        if val is None: return None
        try:
            if isinstance(val, float):
                if math.isnan(val): return None
                return int(val)
            if isinstance(val, int): return val
            s = str(val).strip().split(".")[0]
            return int(s) if s.lstrip("-").isdigit() else None
        except:
            return None

    updated = added = skipped = errors = 0

    try:
        with db_cursor() as (conn, cursor):
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or all(v is None for v in row):
                        continue
                    full_name  = clean_cell(row[1] if len(row) > 1 else None)
                    birth_date = parse_birth_date(row[2] if len(row) > 2 else None)
                    phone      = clean_cell(row[3] if len(row) > 3 else None)
                    hemis      = clean_cell(row[4] if len(row) > 4 else None)
                    uname_raw  = clean_cell(row[5] if len(row) > 5 else None)
                    tg_id_raw  = row[6] if len(row) > 6 else None
                    if not full_name or full_name == "ФИО":
                        skipped += 1
                        continue
                    uname = uname_raw.lstrip("@") if uname_raw else None
                    if not uname: uname = None
                    tg_id = parse_tg_id(tg_id_raw)

                    if tg_id:
                        cursor.execute("SELECT id FROM students WHERE id=%s", (tg_id,))
                        if cursor.fetchone():
                            if uname:
                                cursor.execute(
                                    "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s,username=%s WHERE id=%s",
                                    (full_name, birth_date, phone, hemis, uname, tg_id))
                            else:
                                cursor.execute(
                                    "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s WHERE id=%s",
                                    (full_name, birth_date, phone, hemis, tg_id))
                            updated += 1
                        else:
                            cursor.execute(
                                "INSERT INTO students(id,username,last_active,full_name,birth_date,phone,hemis,started) "
                                "VALUES(%s,%s,%s,%s,%s,%s,%s,0)",
                                (tg_id, uname, now_uz(), full_name, birth_date, phone, hemis))
                            added += 1
                        continue

                    if uname:
                        cursor.execute("SELECT id FROM students WHERE username=%s", (uname,))
                        if cursor.fetchone():
                            cursor.execute(
                                "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s WHERE username=%s",
                                (full_name, birth_date, phone, hemis, uname))
                            updated += 1
                            continue

                    cursor.execute("SELECT id FROM students WHERE full_name=%s", (full_name,))
                    if cursor.fetchone():
                        cursor.execute(
                            "UPDATE students SET birth_date=%s,phone=%s,hemis=%s WHERE full_name=%s",
                            (birth_date, phone, hemis, full_name))
                        updated += 1
                    else:
                        logger.warning(f"Import {row_idx}: TelegramID жоқ, '{full_name}' табылмады.")
                        skipped += 1
                except Exception as e:
                    logger.warning(f"Import row {row_idx}: {e}")
                    try: conn.rollback()
                    except: pass
                    errors += 1
                    continue
            conn.commit()
    except Exception as e:
        logger.error(f"Import критикалық: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Импортта критикалық қате: {e}", reply_markup=excel_submenu())
        return
    finally:
        try: os.remove(path)
        except: pass

    bot.send_message(message.chat.id,
        f"✅ <b>Импорт жуумақланды!</b>\n\n"
        f"🔄 Жаңаланды: <b>{updated}</b>\n"
        f"➕ Қосылды:   <b>{added}</b>\n"
        f"⏭ Өткизилди: <b>{skipped}</b>\n"
        f"❌ Қателер:   <b>{errors}</b>",
        reply_markup=excel_submenu())

# ── МЕНЮ ─────────────────────────────────────────────────────
def main_menu(uid=None):
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("📰 Жаңалықлар", "📚 Сабақ материаллары")
    m.row("📷 Фото/Видео", "📅 Сабақ кестеси")
    m.row("💡 Ұсыныс / Шағым", "📋 Список")
    m.row("📞 Байланыс", "💰 Контракт")
    m.row("📖 Пәнлер", "📊 Сабақ/Ертеңге")
    m.row("🤖 AI Көмекши")
    if uid and is_admin(uid): m.row("👮 Админ панель")
    return m

def back_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.add("⬅️ Артқа")
    return m

def admin_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("👥 Студентлер", "👤 Студент басқарыу")
    m.row("📊 Excel басқарыу", "📊 Барлау басқарыу")
    m.row("📅 Сабақ басқарыу", "❗ Сабақ болмайды")
    m.row("📈 Статистика", "📩 Ус/Ша келген")
    m.row("🗑 Өшириу", "📞 Байланыс басқарыу")
    m.row("💰 Контракт басқарыу", "📖 Пән басқарыу")
    m.row("🔒 Блок басқарыу")
    m.row("⬅️ Артқа")
    return m

def contacts_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("➕ Деканат қосыу", "➕ Муғаллим қосыу")
    m.row("❌ Байланыс өшириу")
    m.row("⬅️ Админге қайтыу")
    return m

def contract_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("💰 Контракт киргизиу", "➕ Төлем қосыу")
    m.row("📋 Барлық контрактлар")
    m.row("⬅️ Админге қайтыу")
    return m

def delete_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("🗑 Материал өшириу", "🗑 Фото/Видео өшириу")
    m.row("🗑 Жаңалық өшириу")
    m.row("⬅️ Админге қайтыу")
    return m

def student_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("➕ Студент қосыу/өзгертиу")
    m.row("❌ Студент өшириу")
    m.row("⬅️ Админге қайтыу")
    return m

def excel_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("📥 Excel жүклеу", "📤 Excel импорт")
    m.row("⬅️ Админге қайтыу")
    return m

def attendance_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("📊 Барлау", "📅 Барлау тарихы")
    m.row("⬅️ Админге қайтыу")
    return m

def schedule_admin_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("➕ Сабақ қосыу", "❌ Сабақ өшириу")
    m.row("⬅️ Админге қайтыу")
    return m

def news_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("✍️ Жазыңыз", "🗂 Архив Жаңалықлар")
    m.row("⬅️ Артқа")
    return m

def materials_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("📥 Мат жүклеңиз", "🗂 Архив материаллар")
    m.row("⬅️ Артқа")
    return m

GALLERY_UPLOAD_BTN = "📤 Жүклеңиз"

def gallery_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row(GALLERY_UPLOAD_BTN, "🎞 S6-DI естелиги")
    m.row("⬅️ Артқа")
    return m

def schedule_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("Понедельник", "Вторник", "Среда")
    m.row("Четверг", "Пятница", "Суббота")
    m.row("Воскресенье")
    m.row("⬅️ Артқа")
    return m

def panler_admin_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("➕ Пән қосыу")
    m.row("🗑 Пән өшириу")
    m.row("⬅️ Админге қайтыу")
    return m

def block_submenu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("🚫 Студентти блоклау")
    m.row("✅ Блоктан шығарыу")
    m.row("📋 Блокланғанлар дизими")
    m.row("⬅️ Админге қайтыу")
    return m

def sabak_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("✅ Бараман", "❌ Себеп бар")
    m.row("⬅️ Артқа")
    return m

def _sebep_file_menu():
    m = types.ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("⏭ Өткизип жибериу")
    m.row("⬅️ Артқа")
    return m

# ── /start ───────────────────────────────────────────────────
@bot.message_handler(commands=["start"])
def start(message):
    uid = message.from_user.id
    username = message.from_user.username or f"user{uid}"
    if is_blocked(uid):
        bot.send_message(uid, "⛔ Сиз блокландыңыз.")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id FROM students WHERE id=%s", (uid,))
        existing = cursor.fetchone()
    if not existing and not is_admin(uid):
        for aid in ADMIN_IDS:
            try:
                fn = message.from_user.first_name or ""
                ln = message.from_user.last_name or ""
                bot.send_message(aid,
                    f"⚠️ <b>Рұхсатсыз кириу!</b>\n👤 {fn} {ln}\n🔗 @{username}\n🆔 <code>{uid}</code>")
            except:
                pass
        bot.send_message(uid,
            "⛔ <b>Кириуге рұхсат жоқ!</b>\nБұл бот тек S6-DI адамлары үшын.\nАдминге хабарласыңыз.")
        return
    with db_cursor() as (conn, cursor):
        cursor.execute(
            "UPDATE students SET username=%s,last_active=%s,started=1 WHERE id=%s",
            (username, now_uz(), uid))
        conn.commit()
    clear_user_state(uid)
    bot.send_message(uid,
        "👋 <b>Хош келдиңиз!</b>\nS6-DI-23 группасы сизлерди көргенимнен қууанышлыман.\nБөлимди таңлаңыз:",
        reply_markup=main_menu(uid))

# ── БЛОК ─────────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "🔒 Блок басқарыу")
@check_access
def block_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "🔒 <b>Блок басқарыу</b>", reply_markup=block_submenu())

@bot.message_handler(func=lambda m: m.text == "🚫 Студентти блоклау")
@check_access
def block_user_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    msg = bot.send_message(message.chat.id,
        "🚫 ID жазыңыз:\n(ямаса <code>ID;себеп</code>)", reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_block_user)

def handle_block_user(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "🔒 Блок басқарыу", reply_markup=block_submenu())
        return
    try:
        parts = [p.strip() for p in message.text.split(";")]
        uid = int(parts[0])
        reason = parts[1] if len(parts) > 1 else "Себеп көрсетилмеген"
    except (ValueError, IndexError):
        msg = bot.send_message(message.chat.id,
            "❌ Дұрыс ID жазыңыз (мысал: 123456789 ямаса 123456789;себеп):",
            reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_block_user)
        return

    if uid in ADMIN_IDS:
        bot.send_message(message.chat.id, "❌ Admin-ді блоклауға болмайды!", reply_markup=block_submenu())
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO blocked_users(user_id,reason) VALUES(%s,%s) "
                "ON CONFLICT(user_id) DO UPDATE SET reason=excluded.reason",
                (uid, reason))
            conn.commit()
        _add_to_blocked_cache(uid)
        bot.send_message(message.chat.id,
            f"✅ <code>{uid}</code> блокланды!\nСебеп: {reason}", reply_markup=block_submenu())
        try:
            bot.send_message(uid, "⛔ Сиз блокландыңыз. Admin-ге хабарласыңыз.")
        except:
            pass
    except Exception as e:
        logger.error(f"handle_block_user DB error: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=block_submenu())

@bot.message_handler(func=lambda m: m.text == "✅ Блоктан шығарыу")
@check_access
def unblock_user_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT user_id,reason FROM blocked_users")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Блокланған жоқ.", reply_markup=block_submenu())
        return
    text = "✅ <b>Блоктан шығарыу — ID жазыңыз:</b>\n\n"
    for r in rows:
        text += f"🆔 <code>{r[0]}</code> — {r[1]}\n"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_unblock_user)

def handle_unblock_user(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "🔒 Блок басқарыу", reply_markup=block_submenu())
        return
    try:
        uid = int(message.text.strip())
    except ValueError:
        msg = bot.send_message(message.chat.id, "❌ Тек сан ID жазыңыз:", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_unblock_user)
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("DELETE FROM blocked_users WHERE user_id=%s", (uid,))
            conn.commit()
        _remove_from_blocked_cache(uid)
        bot.send_message(message.chat.id,
            f"✅ <code>{uid}</code> блоктан шығарылды!", reply_markup=block_submenu())
    except Exception as e:
        logger.error(f"handle_unblock_user: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=block_submenu())

@bot.message_handler(func=lambda m: m.text == "📋 Блокланғанлар дизими")
@check_access
def show_blocked_list(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT user_id,reason,blocked_at FROM blocked_users ORDER BY blocked_at DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Блокланған жоқ.", reply_markup=block_submenu())
        return
    text = f"🔒 <b>Блокланғанлар ({len(rows)}):</b>\n\n"
    for r in rows:
        text += f"🆔 <code>{r[0]}</code>\n📝 {r[1]}\n📅 {r[2]}\n{'─'*20}\n"
    bot.send_message(message.chat.id, text, reply_markup=block_submenu())

# ── НАВИГАЦИЯ ─────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "⬅️ Артқа")
@check_access
def go_back(message):
    uid = message.from_user.id
    mode = get_user_state(uid)
    clear_user_state(uid)
    if mode == "materials":
        bot.send_message(message.chat.id, "📚 Сабақ материаллары", reply_markup=materials_menu())
    elif mode == "gallery":
        bot.send_message(message.chat.id, "📷 Фото/Видео", reply_markup=gallery_menu())
    elif mode == "ai_chat":
        bot.send_message(message.chat.id, "🏠 Бас меню", reply_markup=main_menu(uid))
    elif mode and mode.startswith("variant:") and is_admin(uid):
        bot.send_message(message.chat.id, "📖 Пән басқарыу", reply_markup=panler_admin_submenu())
    elif mode == "sebep_text":
        bot.send_message(message.chat.id, "📊 Сабақ/Ертеңге", reply_markup=sabak_menu())
    elif mode and mode.startswith("sebep_file:"):
        set_user_state(uid, "sebep_text")
        bot.send_message(message.chat.id, "❌ <b>Себебиңизди қайта жазыңыз:</b>", reply_markup=back_menu())
    else:
        bot.send_message(message.chat.id, "🏠 Бас меню", reply_markup=main_menu(uid))

@bot.message_handler(func=lambda m: m.text == "⬅️ Админге қайтыу")
@check_access
def go_back_admin(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫", reply_markup=main_menu(message.from_user.id))
        return
    clear_user_state(message.from_user.id)
    bot.send_message(message.chat.id, "👮 <b>Админ панель</b>", reply_markup=admin_menu())

# ── СПИСОК ────────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📋 Список")
@check_access
def show_student_list(message):
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT full_name,birth_date,phone,hemis FROM students ORDER BY full_name")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Список бос.", reply_markup=main_menu(message.from_user.id))
        return
    HEMIS_URL = "https://student.nukusii.uz/dashboard/login"
    chunks = []
    cur = (f"📋 <b>Студентлер дизими ({len(rows)}):</b>\n"
           f"🎓 <a href='{HEMIS_URL}'>HEMIS Кабинетке кириу →</a>\n\n")
    for i, row in enumerate(rows, 1):
        full_name = row[0] or "—"
        hemis_d = f"<code>{clean_hemis(row[3])}</code>" if clean_hemis(row[3]) else "—"
        phone_d = f"<code>{row[2]}</code>" if row[2] else "—"
        days_left, _ = get_birthday_info(row[1])
        if days_left == 0:
            prefix = "🎂 "
            bd_label = "🎂 <b>Бүгин тууылған күни!!!</b>"
        elif days_left == 1:
            prefix = "🔔 "
            bd_label = "🔔 <b>Ертең тууылған күни!</b>"
        elif days_left is not None and days_left <= 7:
            prefix = "⏳ "
            bd_label = f"⏳ {days_left} күннен кейин тууылған күни"
        else:
            prefix = ""
            bd_label = None
        entry = f"{prefix}{i}. <b>{full_name}</b>\n   📅 {row[1] or '—'}"
        if bd_label: entry += f"\n   {bd_label}"
        entry += f"\n   📞 {phone_d}\n   🎓 HEMIS: {hemis_d}\n{'─'*25}\n"
        if len(cur) + len(entry) > 3800:
            chunks.append(cur)
            cur = ""
        cur += entry
    if cur: chunks.append(cur)
    hemis_mk = types.InlineKeyboardMarkup()
    hemis_mk.add(types.InlineKeyboardButton("🎓 HEMIS Кабинетке кириу", url=HEMIS_URL))
    for i, chunk in enumerate(chunks):
        if i == len(chunks) - 1:
            bot.send_message(message.chat.id, chunk, reply_markup=hemis_mk, disable_web_page_preview=True)
        else:
            bot.send_message(message.chat.id, chunk)
    bot.send_message(message.chat.id, "🏠 Меню:", reply_markup=main_menu(message.from_user.id))

# ── БАЙЛАНЫС ─────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📞 Байланыс")
@check_access
def show_contacts(message):
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT type,name,phone FROM contacts ORDER BY type,name")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Байланыс мағлыуматы жоқ.",
            reply_markup=main_menu(message.from_user.id))
        return
    dekanat = [(r[1], r[2]) for r in rows if r[0] == "dekanat"]
    mugallim = [(r[1], r[2]) for r in rows if r[0] == "mugallim"]
    text = "📞 <b>Байланыс</b>\n\n"
    if dekanat:
        text += "🏛 <b>Деканат:</b>\n"
        for name, phone in dekanat:
            text += f"  👤 {name}\n  📞 <code>{phone}</code>\n\n"
    if mugallim:
        text += "👨‍🏫 <b>Муғаллимлер:</b>\n"
        for name, phone in mugallim:
            text += f"  👤 {name}\n  📞 <code>{phone}</code>\n\n"
    bot.send_message(message.chat.id, text, reply_markup=main_menu(message.from_user.id))

@bot.message_handler(func=lambda m: m.text == "📞 Байланыс басқарыу")
@check_access
def contacts_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "📞 <b>Байланыс басқарыу</b>", reply_markup=contacts_submenu())

@bot.message_handler(func=lambda m: m.text == "➕ Деканат қосыу")
@check_access
def add_dekanat_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    msg = bot.send_message(message.chat.id,
        "🏛 Формат: <code>Аты;Телефон</code>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, lambda m: handle_add_contact(m, "dekanat"))

@bot.message_handler(func=lambda m: m.text == "➕ Муғаллим қосыу")
@check_access
def add_mugallim_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    msg = bot.send_message(message.chat.id,
        "👨‍🏫 Формат: <code>Аты;Телефон</code>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, lambda m: handle_add_contact(m, "mugallim"))

def handle_add_contact(message, contact_type):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📞 Байланыс басқарыу", reply_markup=contacts_submenu())
        return
    parts = [p.strip() for p in message.text.split(";")]
    if len(parts) != 2 or not parts[0] or not parts[1]:
        msg = bot.send_message(message.chat.id,
            "❌ Формат: <code>Аты;Телефон</code>", reply_markup=back_menu())
        bot.register_next_step_handler(msg, lambda m: handle_add_contact(m, contact_type))
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("INSERT INTO contacts(type,name,phone) VALUES(%s,%s,%s)",
                (contact_type, parts[0], parts[1]))
            conn.commit()
        icon = "🏛" if contact_type == "dekanat" else "👨‍🏫"
        bot.send_message(message.chat.id,
            f"✅ {icon} <b>{parts[0]}</b> қосылды!", reply_markup=contacts_submenu())
    except Exception as e:
        logger.error(f"handle_add_contact: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=contacts_submenu())

@bot.message_handler(func=lambda m: m.text == "❌ Байланыс өшириу")
@check_access
def delete_contact_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,type,name,phone FROM contacts ORDER BY type,name")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Байланыслар жоқ.", reply_markup=contacts_submenu())
        return
    text = "❌ <b>Байланыс өшириу — ID жазыңыз:</b>\n\n"
    for r in rows:
        icon = "🏛" if r[1] == "dekanat" else "👨‍🏫"
        text += f"ID:<code>{r[0]}</code> {icon} {r[2]} | 📞 {r[3]}\n"
    text += "\nID ямаса <code>all</code>:"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_delete_contact)

def handle_delete_contact(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📞 Байланыс басқарыу", reply_markup=contacts_submenu())
        return
    try:
        with db_cursor() as (conn, cursor):
            if message.text.strip().lower() == "all":
                cursor.execute("DELETE FROM contacts")
                d = cursor.rowcount
                conn.commit()
                bot.send_message(message.chat.id, f"✅ {d} байланыс өширилди.", reply_markup=contacts_submenu())
            else:
                try:
                    cid = int(message.text.strip())
                except ValueError:
                    msg = bot.send_message(message.chat.id, "❌ ID ямаса all:", reply_markup=back_menu())
                    bot.register_next_step_handler(msg, handle_delete_contact)
                    return
                cursor.execute("SELECT name FROM contacts WHERE id=%s", (cid,))
                row = cursor.fetchone()
                if not row:
                    bot.send_message(message.chat.id, "⚠️ Табылмады.", reply_markup=contacts_submenu())
                    return
                cursor.execute("DELETE FROM contacts WHERE id=%s", (cid,))
                conn.commit()
                bot.send_message(message.chat.id,
                    f"✅ <b>{row[0]}</b> өширилди.", reply_markup=contacts_submenu())
    except Exception as e:
        logger.error(f"handle_delete_contact: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=contacts_submenu())

# ── КОНТРАКТ ──────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "💰 Контракт")
@check_access
def show_contract_user(message):
    uid = message.from_user.id
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT total_amount,note FROM contracts WHERE student_id=%s", (uid,))
        contract = cursor.fetchone()
        if not contract:
            bot.send_message(message.chat.id,
                "📭 Контрактыңыз орнатылмаған.\nАдминге хабарласыңыз.",
                reply_markup=main_menu(uid))
            return
        total = contract[0]
        note = contract[1] or ""
        cursor.execute(
            "SELECT COALESCE(SUM(amount),0) FROM contract_payments WHERE student_id=%s", (uid,))
        paid = float(cursor.fetchone()[0])
        remaining = total - paid
        cursor.execute(
            "SELECT amount,date,note FROM contract_payments WHERE student_id=%s ORDER BY date DESC",
            (uid,))
        payments = cursor.fetchall()
        # FIX: N+1 query-ден арылу — бір JOIN query-мен барлығын алу
        cursor.execute("""
            SELECT s.full_name, s.id, c.total_amount,
                COALESCE(SUM(p.amount), 0) as paid
            FROM students s
            JOIN contracts c ON c.student_id = s.id
            LEFT JOIN contract_payments p ON p.student_id = s.id
            WHERE s.full_name IS NOT NULL
            GROUP BY s.full_name, s.id, c.total_amount
            ORDER BY s.full_name""")
        all_contracts = cursor.fetchall()

    percent = int((paid / total) * 100) if total > 0 else 0
    bar = "🟩" * (percent // 10) + "⬜" * (10 - percent // 10)
    text = f"💰 <b>Мениң контрактым</b>\n{'─'*30}\n"
    if note: text += f"📝 {note}\n"
    text += (f"\n💵 Улыума: <b>{total:,.0f} сум</b>\n✅ Төленди: <b>{paid:,.0f} сум</b>\n"
             f"⏳ Қалды: <b>{remaining:,.0f} сум</b>\n{bar} <b>{percent}%</b>\n{'─'*30}\n")
    if payments:
        text += "\n📜 <b>Төлем тарихы:</b>\n"
        for p in payments:
            p_note = f" — {p[2]}" if p[2] else ""
            text += f"  ✅ {date_to_ru(p[1])} | <b>{p[0]:,.0f} сум</b>{p_note}\n"
    else:
        text += "\n📭 Төлем тарихы жоқ.\n"
    if all_contracts:
        text += f"\n{'─'*30}\n📋 <b>Группаның жағдайы ({len(all_contracts)} студент):</b>\n\n"
        for r in all_contracts:
            s_remain = r[2] - float(r[3])
            s_pct = int((float(r[3]) / r[2]) * 100) if r[2] > 0 else 0
            s_bar = "🟩" * (s_pct // 10) + "⬜" * (10 - s_pct // 10)
            me = " 👈 <i>сиз</i>" if r[1] == uid else ""
            if s_remain <= 0:
                text += f"✅ <b>{r[0]}</b>{me}\n   {s_bar} <b>100%</b> — Толық төленди\n\n"
            else:
                text += f"⏳ <b>{r[0]}</b>{me}\n   {s_bar} <b>{s_pct}%</b>\n   Қалды: <b>{s_remain:,.0f} сум</b>\n\n"
    send_long_message(message.chat.id, text, reply_markup=main_menu(uid))

@bot.message_handler(func=lambda m: m.text == "💰 Контракт басқарыу")
@check_access
def contract_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "💰 <b>Контракт басқарыу</b>", reply_markup=contract_submenu())

@bot.message_handler(func=lambda m: m.text == "💰 Контракт киргизиу")
@check_access
def contract_set_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,full_name FROM students WHERE full_name IS NOT NULL ORDER BY full_name")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Студентлер жоқ.", reply_markup=contract_submenu())
        return
    text = "💰 <b>Контракт киргизиу:</b>\nФормат: <code>TelegramID;Сумма;Ескертіу</code>\n\n📋 <b>Студентлер:</b>\n"
    for r in rows:
        text += f"🆔 <code>{r[0]}</code> — {r[1]}\n"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_contract_set)

def handle_contract_set(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "💰 Контракт басқарыу", reply_markup=contract_submenu())
        return
    parts = [p.strip() for p in message.text.split(";")]
    if len(parts) < 2 or not parts[0].lstrip("-").isdigit():
        msg = bot.send_message(message.chat.id,
            "❌ Формат: <code>TelegramID;Сумма;Ескертиу</code>", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_contract_set)
        return
    try:
        sid = int(parts[0])
        amount = float(parts[1].replace(" ", "").replace(",", ""))
        note = parts[2] if len(parts) > 2 else ""
    except ValueError as e:
        msg = bot.send_message(message.chat.id, f"❌ Формат қатеси: {e}", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_contract_set)
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("SELECT full_name FROM students WHERE id=%s", (sid,))
            row = cursor.fetchone()
            if not row:
                msg = bot.send_message(message.chat.id, "⚠️ Студент табылмады.", reply_markup=back_menu())
                bot.register_next_step_handler(msg, handle_contract_set)
                return
            cursor.execute(
                "INSERT INTO contracts(student_id,total_amount,note) VALUES(%s,%s,%s) "
                "ON CONFLICT(student_id) DO UPDATE SET total_amount=excluded.total_amount,note=excluded.note",
                (sid, amount, note))
            conn.commit()
            name = row[0]
        bot.send_message(message.chat.id,
            f"✅ <b>{name}</b>\n💵 Контракт: <b>{amount:,.0f} сум</b>",
            reply_markup=contract_submenu())
        try:
            bot.send_message(sid,
                f"💰 <b>Контрактыңыз киргизилди!</b>\n💵 Улыума: <b>{amount:,.0f} сум</b>"
                + (f"\n📝 {note}" if note else ""))
        except:
            pass
    except Exception as e:
        logger.error(f"handle_contract_set: {e}", exc_info=True)
        msg = bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_contract_set)

@bot.message_handler(func=lambda m: m.text == "➕ Төлем қосыу")
@check_access
def payment_add_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("""
            SELECT s.id, s.full_name, c.total_amount,
                COALESCE(SUM(p.amount), 0) as paid
            FROM students s
            JOIN contracts c ON c.student_id = s.id
            LEFT JOIN contract_payments p ON p.student_id = s.id
            WHERE s.full_name IS NOT NULL
            GROUP BY s.id, s.full_name, c.total_amount
            ORDER BY s.full_name""")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id,
            "📭 Контракт киргизилген студент жоқ.", reply_markup=contract_submenu())
        return
    text = "➕ <b>Төлем қосыу:</b>\nФормат: <code>TelegramID;Сумма;Ескертиу</code>\n\n📋 <b>Контрактлар:</b>\n"
    for r in rows:
        rem = r[2] - float(r[3])
        text += f"{'✅' if rem <= 0 else '⏳'} <code>{r[0]}</code> — {r[1]}\n   Қалды: <b>{rem:,.0f} сум</b>\n"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_payment_add)

def handle_payment_add(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "💰 Контракт басқарыу", reply_markup=contract_submenu())
        return
    parts = [p.strip() for p in message.text.split(";")]
    if len(parts) < 2 or not parts[0].lstrip("-").isdigit():
        msg = bot.send_message(message.chat.id,
            "❌ Формат: <code>TelegramID;Сумма;Ескертиу</code>", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_payment_add)
        return
    try:
        sid = int(parts[0])
        amount = float(parts[1].replace(" ", "").replace(",", ""))
        note = parts[2] if len(parts) > 2 else ""
    except ValueError as e:
        msg = bot.send_message(message.chat.id, f"❌ Формат қатесі: {e}", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_payment_add)
        return
    ds = now_uz().strftime("%Y-%m-%d")
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("SELECT full_name FROM students WHERE id=%s", (sid,))
            sr = cursor.fetchone()
            cursor.execute("SELECT total_amount FROM contracts WHERE student_id=%s", (sid,))
            cr = cursor.fetchone()
            if not sr or not cr:
                msg = bot.send_message(message.chat.id,
                    "⚠️ Студент ямаса контракт табылмады.", reply_markup=back_menu())
                bot.register_next_step_handler(msg, handle_payment_add)
                return
            cursor.execute(
                "INSERT INTO contract_payments(student_id,amount,date,note) VALUES(%s,%s,%s,%s)",
                (sid, amount, ds, note))
            cursor.execute(
                "SELECT COALESCE(SUM(amount),0) FROM contract_payments WHERE student_id=%s", (sid,))
            paid = float(cursor.fetchone()[0])
            total = cr[0]
            rem = total - paid
            conn.commit()
            name = sr[0]
        bot.send_message(message.chat.id,
            f"✅ Төлем қосылды!\n👤 <b>{name}</b>\n💵 {amount:,.0f} сум\n⏳ Қалды: <b>{rem:,.0f} сум</b>",
            reply_markup=contract_submenu())
        try:
            pct = int((paid / total) * 100) if total > 0 else 0
            bar = "🟩" * (pct // 10) + "⬜" * (10 - pct // 10)
            bot.send_message(sid,
                f"💰 <b>Төлем қабылланды!</b>\n📅 {date_to_ru(ds)}\n{'─'*25}\n"
                f"✅ Төленди: <b>{amount:,.0f} сум</b>\n⏳ Қалды: <b>{rem:,.0f} сум</b>\n\n{bar} {pct}%")
        except:
            pass
    except Exception as e:
        logger.error(f"handle_payment_add: {e}", exc_info=True)
        msg = bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_payment_add)

@bot.message_handler(func=lambda m: m.text == "📋 Барлық контрактлар")
@check_access
def show_all_contracts(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("""
            SELECT s.full_name, c.total_amount,
                COALESCE(SUM(p.amount), 0) as paid
            FROM students s
            JOIN contracts c ON c.student_id = s.id
            LEFT JOIN contract_payments p ON p.student_id = s.id
            WHERE s.full_name IS NOT NULL
            GROUP BY s.full_name, c.total_amount
            ORDER BY s.full_name""")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Контрактлар жоқ.", reply_markup=contract_submenu())
        return
    ts = sum(r[1] for r in rows)
    ps = sum(float(r[2]) for r in rows)
    text = (f"📋 <b>Барлық контрактлар ({len(rows)}):</b>\n"
            f"💵 Улыума: <b>{ts:,.0f}</b>\n✅ Түскен: <b>{ps:,.0f}</b>\n"
            f"⏳ Қалды: <b>{ts-ps:,.0f} сум</b>\n{'─'*30}\n\n")
    for r in rows:
        rem = r[1] - float(r[2])
        status = "✅ Толық" if rem <= 0 else f"⏳ Қалды: {rem:,.0f}"
        text += f"👤 <b>{r[0]}</b>\n   💵 {r[1]:,.0f} | ✅ {float(r[2]):,.0f} | {status}\n\n"
    send_long_message(message.chat.id, text, reply_markup=contract_submenu())

# ── ЖАҢАЛЫҚЛАР ───────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📰 Жаңалықлар")
@check_access
def show_news_menu(message):
    bot.send_message(message.chat.id, "📰 <b>Жаңалықлар бөлими</b>", reply_markup=news_menu())

@bot.message_handler(func=lambda m: m.text == "✍️ Жазыңыз")
@check_access
def write_news(message):
    msg = bot.send_message(message.chat.id, "✍️ <b>Жаңалығыңызды жазыңыз:</b>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_user_news)

def handle_user_news(message):
    if not user_step_check(message):
        return
    if not message.text:
        msg = bot.send_message(message.chat.id, "✍️ Тек текст жибериңиз:", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_user_news)
        return
    if message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📰 Жаңалықлар", reply_markup=news_menu())
        return
    if len(message.text) > 2000:
        msg = bot.send_message(message.chat.id,
            "❌ Текст дым ұзын (макс 2000 таңба).", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_user_news)
        return
    uid = message.from_user.id
    username = message.from_user.username or f"user{uid}"
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO user_news(content,author_id,author_username) VALUES(%s,%s,%s)",
                (message.text, uid, username))
            conn.commit()
        send_to_students(
            text=f"📰 <b>Таза хабарлама!</b>\n\n👤 <b>@{username}</b>:\n\n{message.text}",
            exclude_id=uid)
        bot.send_message(message.chat.id, "✅ Жиберилди!", reply_markup=news_menu())
    except Exception as e:
        logger.error(f"handle_user_news: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Қате: {e}", reply_markup=news_menu())

@bot.message_handler(func=lambda m: m.text == "🗂 Архив Жаңалықлар")
@check_access
def show_news_archive(message):
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT content,author_username,date FROM user_news ORDER BY date DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Архив бос.", reply_markup=news_menu())
        return
    chunks = []
    cur = "🗂 <b>Архив жаңалықлар:</b>\n\n"
    for r in rows:
        entry = f"👤 <b>@{r[1]}</b>\n📌 {r[0]}\n🕐 {r[2]}\n{'─'*25}\n"
        if len(cur) + len(entry) > 3800:
            chunks.append(cur)
            cur = ""
        cur += entry
    if cur: chunks.append(cur)
    for i, chunk in enumerate(chunks):
        bot.send_message(message.chat.id, chunk,
            reply_markup=news_menu() if i == len(chunks) - 1 else None)

# ── МАТЕРИАЛДАР ───────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📚 Сабақ материаллары")
@check_access
def show_materials_menu(message):
    bot.send_message(message.chat.id, "📚 <b>Сабақ материаллары</b>", reply_markup=materials_menu())

@bot.message_handler(func=lambda m: m.text == "📥 Мат жүклеңиз")
@check_access
def upload_material_start(message):
    set_user_state(message.from_user.id, "materials")
    bot.send_message(message.chat.id,
        "📥 <b>Файл ямаса фото жибериңиз:</b>\nТайын болғанда <b>⬅️ Артқа</b> басыңыз.",
        reply_markup=back_menu())

@bot.message_handler(content_types=["document"],
    func=lambda m: get_user_state(m.from_user.id) == "materials")
@check_access
def handle_upload_document(message):
    if is_already_processed(message.message_id): return
    uid = message.from_user.id
    username = message.from_user.username or f"user{uid}"
    file_id = message.document.file_id
    file_name = message.document.file_name or "Файл"
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO materials(file_id,file_type,uploader_id,uploader_username) VALUES(%s,%s,%s,%s)",
                (file_id, "document", uid, username))
            conn.commit()
        send_to_students(file_id=file_id, file_type="document",
            text=f"📚 <b>Таза материал!</b>\n👤 @{username}\n📎 {file_name}", exclude_id=uid)
        send_saved_once(message.chat.id, uid)
    except Exception as e:
        logger.error(f"handle_upload_document: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Қате: {e}")

@bot.message_handler(content_types=["photo"],
    func=lambda m: get_user_state(m.from_user.id) == "materials")
@check_access
def handle_upload_photo_mat(message):
    if is_already_processed(message.message_id): return
    uid = message.from_user.id
    username = message.from_user.username or f"user{uid}"
    file_id = message.photo[-1].file_id
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO materials(file_id,file_type,uploader_id,uploader_username) VALUES(%s,%s,%s,%s)",
                (file_id, "photo", uid, username))
            conn.commit()
        send_to_students(file_id=file_id, file_type="photo",
            text=f"📚 <b>Таза материал!</b>\n👤 @{username}", exclude_id=uid)
        send_saved_once(message.chat.id, uid)
    except Exception as e:
        logger.error(f"handle_upload_photo_mat: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Қате: {e}")

@bot.message_handler(content_types=["video", "audio", "voice", "sticker"],
    func=lambda m: get_user_state(m.from_user.id) == "materials")
@check_access
def handle_upload_wrong_materials(message):
    bot.send_message(message.chat.id, "⚠️ Тек файл ямаса фото!")

@bot.message_handler(func=lambda m: m.text == "🗂 Архив материаллар")
@check_access
def show_materials_archive(message):
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT file_id,file_type,date,uploader_username FROM materials ORDER BY date DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Архив бос.", reply_markup=materials_menu())
        return
    bot.send_message(message.chat.id, f"🗂 <b>Барлығы: {len(rows)}</b>\n\nЖүклениуде...")
    for r in rows:
        uname = f"@{r[3]}" if r[3] else "Белгисиз"
        cap = f"👤 {uname}\n🕐 {r[2]}"
        try:
            if r[1] == "document": bot.send_document(message.chat.id, r[0], caption=cap)
            elif r[1] == "photo": bot.send_photo(message.chat.id, r[0], caption=cap)
        except:
            continue
    bot.send_message(message.chat.id, "✅ Тайын.", reply_markup=materials_menu())

# ── ГАЛЕРЕЯ ───────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📷 Фото/Видео")
@check_access
def show_gallery_menu(message):
    bot.send_message(message.chat.id, "📷 <b>Фото/Видео бөлими</b>", reply_markup=gallery_menu())

@bot.message_handler(func=lambda m: m.text == GALLERY_UPLOAD_BTN)
@check_access
def gallery_upload_start(message):
    set_user_state(message.from_user.id, "gallery")
    bot.send_message(message.chat.id,
        "📤 <b>Фото ямаса видео жибериңиз:</b>\nТайын болғанда <b>⬅️ Артқа</b> басыңыз.",
        reply_markup=back_menu())

@bot.message_handler(content_types=["photo"],
    func=lambda m: get_user_state(m.from_user.id) == "gallery")
@check_access
def handle_gallery_photo(message):
    if is_already_processed(message.message_id): return
    uid = message.from_user.id
    username = message.from_user.username or f"user{uid}"
    file_id = message.photo[-1].file_id
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO gallery(file_id,file_type,uploader_id,uploader_username) VALUES(%s,%s,%s,%s)",
                (file_id, "photo", uid, username))
            conn.commit()
        send_to_students(file_id=file_id, file_type="photo",
            text=f"🎞 <b>S6-DI естелиги!</b>\n👤 @{username}", exclude_id=uid)
        send_saved_once(message.chat.id, uid)
    except Exception as e:
        logger.error(f"handle_gallery_photo: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Қате: {e}")

@bot.message_handler(content_types=["video"],
    func=lambda m: get_user_state(m.from_user.id) == "gallery")
@check_access
def handle_gallery_video(message):
    if is_already_processed(message.message_id): return
    uid = message.from_user.id
    username = message.from_user.username or f"user{uid}"
    file_id = message.video.file_id
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO gallery(file_id,file_type,uploader_id,uploader_username) VALUES(%s,%s,%s,%s)",
                (file_id, "video", uid, username))
            conn.commit()
        send_to_students(file_id=file_id, file_type="video",
            text=f"🎞 <b>S6-DI естелиги!</b>\n👤 @{username}", exclude_id=uid)
        send_saved_once(message.chat.id, uid)
    except Exception as e:
        logger.error(f"handle_gallery_video: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Қате: {e}")

@bot.message_handler(content_types=["document", "audio", "voice", "sticker"],
    func=lambda m: get_user_state(m.from_user.id) == "gallery")
@check_access
def handle_upload_wrong_gallery(message):
    bot.send_message(message.chat.id, "⚠️ Тек фото ямаса видео!")

@bot.message_handler(func=lambda m: m.text == "🎞 S6-DI естелиги")
@check_access
def show_gallery_view(message):
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT file_id,file_type,date,uploader_username FROM gallery ORDER BY date DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Галерея бос.", reply_markup=gallery_menu())
        return
    bot.send_message(message.chat.id, f"🎞 <b>Барлығы: {len(rows)}</b>\n\nЖүкленуде...")
    for r in rows:
        uname = f"@{r[3]}" if r[3] else "Белгісіз"
        cap = f"👤 {uname}\n📅 {r[2]}"
        try:
            if r[1] == "photo": bot.send_photo(message.chat.id, r[0], caption=cap)
            elif r[1] == "video": bot.send_video(message.chat.id, r[0], caption=cap)
        except:
            continue
    bot.send_message(message.chat.id, "✅ Тайын.", reply_markup=gallery_menu())

# ── САБАҚ КЕСТЕСІ ─────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📅 Сабақ кестеси")
@check_access
def show_schedule_menu(message):
    bot.send_message(message.chat.id, "📅 <b>Сабақ кестесі</b>\nКүнди таңлаңыз:", reply_markup=schedule_menu())

@bot.message_handler(func=lambda m: m.text in DAYS_RU)
@check_access
def show_day_schedule(message):
    day = message.text
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT subject,time FROM schedule WHERE day=%s ORDER BY time", (day,))
        rows = cursor.fetchall()
    today_ru = DAYS_EN_TO_RU.get(now_uz().strftime("%A"), "")
    today_mark = " 📌 <i>(бүгин)</i>" if day == today_ru else ""
    if not rows:
        bot.send_message(message.chat.id,
            f"📭 <b>{day}{today_mark}</b>\n\nСабақ жоқ.", reply_markup=schedule_menu())
        return
    text = f"📅 <b>{day}{today_mark}</b>\n\n"
    for i, r in enumerate(rows, 1):
        text += f"{i}-пара 🕐 <b>{r[1]}</b> — {r[0]}\n"
    bot.send_message(message.chat.id, text, reply_markup=schedule_menu())

# ── ҰСЫНЫС ───────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "💡 Ұсыныс / Шағым")
@check_access
def suggestion_start(message):
    msg = bot.send_message(message.chat.id,
        "💡 <b>Ұсыныс ямаса шағымыңызды жазыңыз:</b>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_suggestion)

def handle_suggestion(message):
    if not user_step_check(message):
        return
    if not message.text:
        msg = bot.send_message(message.chat.id, "✍️ Текст жибериңиз:", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_suggestion)
        return
    if message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "🏠 Бас меню", reply_markup=main_menu(message.from_user.id))
        return
    if len(message.text) > 1000:
        msg = bot.send_message(message.chat.id,
            "❌ Текст дым ұзын (макс 1000 таңба).", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_suggestion)
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("INSERT INTO suggestions(content,user_id) VALUES(%s,%s)",
                (message.text, message.from_user.id))
            conn.commit()
        bot.send_message(message.chat.id, "✅ Жиберилди! Рахмет!", reply_markup=main_menu(message.from_user.id))
        for aid in ADMIN_IDS:
            try:
                fn = message.from_user.first_name or ""
                ln = message.from_user.last_name or ""
                un = f"@{message.from_user.username}" if message.from_user.username else "username жоқ"
                bot.send_message(aid,
                    f"💡 <b>Таза ұсыныс/шағым:</b>\n\n{message.text}\n\n"
                    f"👤 {fn} {ln}\n🔗 {un}\n🆔 <code>{message.from_user.id}</code>")
            except:
                pass
    except Exception as e:
        logger.error(f"handle_suggestion: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Қате: {e}", reply_markup=main_menu(message.from_user.id))

# ── АДМИН ПАНЕЛЬ ─────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "👮 Админ панель")
@check_access
def admin_panel(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫 Сиз админ емессиз!")
        return
    bot.send_message(message.chat.id, "👮 <b>Админ панель</b>", reply_markup=admin_menu())

@bot.message_handler(func=lambda m: m.text == "👤 Студент басқарыу")
@check_access
def student_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "👤 <b>Студент басқарыу</b>", reply_markup=student_submenu())

@bot.message_handler(func=lambda m: m.text == "➕ Студент қосыу/өзгертиу")
@check_access
def student_add_or_edit_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,full_name,username FROM students ORDER BY full_name")
        rows = cursor.fetchall()
    header = ("➕ <b>Студент қосыу / Өзгертіу:</b>\n\n🆕 <b>Таза қосыу:</b>\n"
              "<code>жаңа;ФИО;Туылған күни;Тел;HEMIS;TelegramID</code>\n"
              "📌 Мысал: <code>таза;Иванов Иван;2000-01-01;+998901234567;S12345678;123456789</code>\n\n"
              "✏️ <b>Өзгертиу:</b> студент ID-ін жазыңыз\n" + "─" * 30 + "\n")
    if not rows:
        msg = bot.send_message(message.chat.id, header + "📭 Студентлер жоқ.", reply_markup=back_menu())
        bot.register_next_step_handler(msg, student_add_or_edit)
        return
    chunks = []
    cur = header
    for i, r in enumerate(rows, 1):
        line = f"{i}. 👤 <b>{r[1] or '—'}</b>\n    🆔 <code>{r[0]}</code> | {'@'+r[2] if r[2] else 'username жоқ'}\n"
        if len(cur) + len(line) > 3800:
            chunks.append(cur)
            cur = ""
        cur += line
    cur += "─" * 30 + "\n⬇️ <b>ID жазыңыз ямаса таза студент форматын жибериңиз:</b>"
    chunks.append(cur)
    for chunk in chunks[:-1]:
        bot.send_message(message.chat.id, chunk)
    msg = bot.send_message(message.chat.id, chunks[-1], reply_markup=back_menu())
    bot.register_next_step_handler(msg, student_add_or_edit)

@bot.message_handler(func=lambda m: m.text == "❌ Студент өшириу")
@check_access
def student_delete_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,full_name,username FROM students ORDER BY full_name")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Студентлер жоқ.", reply_markup=student_submenu())
        return
    text = "❌ <b>Студент өшириу — ID жазыңыз:</b>\n\n"
    for r in rows:
        text += f"ID:<code>{r[0]}</code> — {r[1] or '—'} (@{r[2] or '—'})\n"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, delete_student)

def student_add_or_edit(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "👤 Студент басқарыу", reply_markup=student_submenu())
        return
    # FIX: "таза" -> "таза" деп өзгертилди (ямаса екеуин де қабыллайды)
    if message.text.strip().lower().startswith(("таза;", "таза;")):
        parts = [p.strip() for p in message.text.split(";")]
        if len(parts) < 6 or not parts[1] or not parts[5]:
            msg = bot.send_message(message.chat.id,
                "❌ Формат:\n<code>жаңа;ФИО;Тууылған күни;Тел;HEMIS;TelegramID</code>",
                reply_markup=back_menu())
            bot.register_next_step_handler(msg, student_add_or_edit)
            return
        if not parts[5].lstrip("-").isdigit():
            msg = bot.send_message(message.chat.id,
                "❌ TelegramID тек сан болуы керек!", reply_markup=back_menu())
            bot.register_next_step_handler(msg, student_add_or_edit)
            return
        fn = parts[1]
        bd = parts[2] if len(parts) > 2 else ""
        ph = parts[3] if len(parts) > 3 else ""
        hm = parts[4] if len(parts) > 4 else ""
        tg_id = int(parts[5])
        try:
            with db_cursor() as (conn, cursor):
                cursor.execute("SELECT id,full_name FROM students WHERE id=%s", (tg_id,))
                ex = cursor.fetchone()
                if ex:
                    bot.send_message(message.chat.id,
                        f"⚠️ Бұл ID бұрыннан бар!\n👤 {ex[1] or '—'}\n\nӨзгертиу үшын ID жазыңыз: <code>{tg_id}</code>",
                        reply_markup=back_menu())
                    bot.register_next_step_handler(message, student_add_or_edit)
                    return
                cursor.execute(
                    "INSERT INTO students(id,username,last_active,full_name,birth_date,phone,hemis) "
                    "VALUES(%s,%s,%s,%s,%s,%s,%s)",
                    (tg_id, None, now_uz(), fn, bd, ph, hm))
                conn.commit()
            bot.send_message(message.chat.id,
                f"✅ <b>{fn}</b> қосылды!\n🆔 <code>{tg_id}</code>\n🎓 HEMIS: {hm}\n\n"
                "📌 Студент ботқа /start берсин.",
                reply_markup=student_submenu())
        except Exception as e:
            logger.error(f"student_add_or_edit insert: {e}", exc_info=True)
            bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=student_submenu())
        return
    try:
        sid = int(message.text.strip())
    except ValueError:
        msg = bot.send_message(message.chat.id,
            "❌ ID ямаса <code>таза;ФИО;Күн;Тел;HEMIS;TelegramID</code>",
            reply_markup=back_menu())
        bot.register_next_step_handler(msg, student_add_or_edit)
        return
    try:
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT id,full_name,birth_date,phone,hemis FROM students WHERE id=%s", (sid,))
            row = cursor.fetchone()
        if not row:
            msg = bot.send_message(message.chat.id, "⚠️ ID табылмады:", reply_markup=back_menu())
            bot.register_next_step_handler(msg, student_add_or_edit)
            return
        sid, fname, bdate, phone, hemis = row
        text = (f"✏️ <b>Студент:</b>\n"
                f"👤 {fname or '—'} | 📅 {bdate or '—'} | 📞 {phone or '—'} | 🎓 {hemis or '—'}\n\n"
                "<code>ФИО;Күн;Тел;HEMIS</code>\nӨзгертпей <b>—</b> жазыңыз.")
        msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
        bot.register_next_step_handler(msg,
            lambda m: student_edit_save(m, sid, fname, bdate, phone, hemis))
    except Exception as e:
        logger.error(f"student_add_or_edit fetch: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=student_submenu())

def student_edit_save(message, sid, old_fn, old_bd, old_ph, old_hm):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "👤 Студент басқарыу", reply_markup=student_submenu())
        return
    try:
        parts = [p.strip() for p in message.text.split(";")]
        if len(parts) != 4: raise ValueError("4 бөлик болыуы керек")
        nf = parts[0] if parts[0] != "—" else old_fn
        nb = parts[1] if parts[1] != "—" else old_bd
        np_ = parts[2] if parts[2] != "—" else old_ph
        nh = parts[3] if parts[3] != "—" else old_hm
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s WHERE id=%s",
                (nf, nb, np_, nh, sid))
            conn.commit()
        bot.send_message(message.chat.id,
            f"✅ <b>{nf}</b> тазаланды!\n🎓 HEMIS: {nh}", reply_markup=student_submenu())
    except Exception as e:
        msg = bot.send_message(message.chat.id,
            f"❌ <code>ФИО;Күн;Тел;HEMIS</code> ({e})", reply_markup=back_menu())
        bot.register_next_step_handler(msg,
            lambda m: student_edit_save(m, sid, old_fn, old_bd, old_ph, old_hm))

def delete_student(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "👤 Студент басқарыу", reply_markup=student_submenu())
        return
    try:
        sid = int(message.text.strip())
    except ValueError:
        msg = bot.send_message(message.chat.id, "❌ Тек сан ID жазыңыз:", reply_markup=back_menu())
        bot.register_next_step_handler(msg, delete_student)
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute("SELECT full_name FROM students WHERE id=%s", (sid,))
            row = cursor.fetchone()
            if not row:
                bot.send_message(message.chat.id, "⚠️ Табылмады.", reply_markup=student_submenu())
                return
            cursor.execute("DELETE FROM students WHERE id=%s", (sid,))
            cursor.execute("DELETE FROM attendance WHERE student_id=%s", (sid,))
            # FIX: Студент өширилгенде байланыслы барлық данныйларды тазалау
            cursor.execute("DELETE FROM contracts WHERE student_id=%s", (sid,))
            cursor.execute("DELETE FROM contract_payments WHERE student_id=%s", (sid,))
            cursor.execute("DELETE FROM blocked_users WHERE user_id=%s", (sid,))
            cursor.execute("DELETE FROM user_states WHERE user_id=%s", (sid,))
            conn.commit()
        # State cache-тен де өшіру
        with _state_cache_lock:
            _state_cache.pop(sid, None)
        bot.send_message(message.chat.id, f"✅ <b>{row[0]}</b> өширилди.", reply_markup=student_submenu())
    except Exception as e:
        logger.error(f"delete_student: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=student_submenu())

# ── EXCEL HANDLERS ────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📊 Excel басқарыу")
@check_access
def excel_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "📊 <b>Excel басқарыу</b>", reply_markup=excel_submenu())

@bot.message_handler(func=lambda m: m.text == "📥 Excel жүклеу")
@check_access
def excel_download(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    _excel_download_impl(message)

@bot.message_handler(func=lambda m: m.text == "📤 Excel импорт")
@check_access
def excel_import_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    msg = bot.send_message(message.chat.id,
        "📤 <b>Excel файлды жибериңиз (.xlsx):</b>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_excel_import)

def handle_excel_import(message):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📊 Excel басқарыу", reply_markup=excel_submenu())
        return
    _excel_import_impl(message)

# ── САБАҚ БАСҚАРЫУ ────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📅 Сабақ басқарыу")
@check_access
def schedule_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "📅 <b>Сабақ басқарыу</b>", reply_markup=schedule_admin_submenu())

@bot.message_handler(func=lambda m: m.text == "➕ Сабақ қосыу")
@check_access
def schedule_add_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    msg = bot.send_message(message.chat.id,
        "📝 Формат: <code>Понедельник;Математика;09:00</code>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, add_lesson)

@bot.message_handler(func=lambda m: m.text == "❌ Сабақ өшириу")
@check_access
def schedule_delete_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,day,subject,time FROM schedule ORDER BY day,time")
        lessons = cursor.fetchall()
    if not lessons:
        bot.send_message(message.chat.id, "📭 Кесте бос.", reply_markup=schedule_admin_submenu())
        return
    text = "📋 <b>Барлық сабақлар:</b>\n\n"
    for r in lessons:
        text += f"ID:{r[0]} | {r[1]} | {r[2]} | {r[3]}\n"
    text += "\n<code>Күн;Уақыт</code> форматында жазыңыз:"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, delete_lesson)

def add_lesson(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📅 Сабақ басқарыу", reply_markup=schedule_admin_submenu())
        return
    try:
        parts = [p.strip() for p in message.text.split(";")]
        if len(parts) != 3 or not all(parts): raise ValueError
        day, subject, time_ = parts
        if day not in DAYS_RU: raise ValueError(f"Күн дұрыс емес: {day}")
        with db_cursor() as (conn, cursor):
            cursor.execute("INSERT INTO schedule(day,subject,time) VALUES(%s,%s,%s)",
                (day, subject, time_))
            conn.commit()
        bot.send_message(message.chat.id,
            f"✅ <b>{day} | {subject} | {time_}</b> қосылды!",
            reply_markup=schedule_admin_submenu())
    except ValueError as e:
        msg = bot.send_message(message.chat.id,
            f"❌ <code>Понедельник;Математика;09:00</code>\n({e})", reply_markup=back_menu())
        bot.register_next_step_handler(msg, add_lesson)
    except Exception as e:
        logger.error(f"add_lesson: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=schedule_admin_submenu())

def delete_lesson(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📅 Сабақ басқарыу", reply_markup=schedule_admin_submenu())
        return
    try:
        parts = [p.strip() for p in message.text.split(";")]
        if len(parts) != 2 or not all(parts): raise ValueError
        day, time_ = parts
        with db_cursor() as (conn, cursor):
            cursor.execute("DELETE FROM schedule WHERE day=%s AND time=%s", (day, time_))
            d = cursor.rowcount
            conn.commit()
        if d:
            bot.send_message(message.chat.id,
                f"✅ Өширилди: {day} — {time_}", reply_markup=schedule_admin_submenu())
        else:
            bot.send_message(message.chat.id, "⚠️ Табылмады.", reply_markup=schedule_admin_submenu())
    except ValueError:
        msg = bot.send_message(message.chat.id,
            "❌ <code>Понедельник;09:00</code>", reply_markup=back_menu())
        bot.register_next_step_handler(msg, delete_lesson)
    except Exception as e:
        logger.error(f"delete_lesson: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=schedule_admin_submenu())

# ── БАРЛАУ ────────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📊 Барлау басқарыу")
@check_access
def attendance_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "📊 <b>Барлау басқарыу</b>", reply_markup=attendance_submenu())

@bot.message_handler(func=lambda m: m.text == "📊 Барлау")
@check_access
def start_attendance(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    today = DAYS_EN_TO_RU.get(now_uz().strftime("%A"), "")
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT subject,time FROM schedule WHERE day=%s ORDER BY time", (today,))
        lessons = cursor.fetchall()
    if not lessons:
        bot.send_message(message.chat.id,
            f"📭 Бүгин ({today}) сабақ жоқ.", reply_markup=attendance_submenu())
        return
    markup = types.InlineKeyboardMarkup()
    for i, (subject, time_) in enumerate(lessons, 1):
        markup.add(types.InlineKeyboardButton(
            text=f"{i}-пара: {subject} ({time_})",
            callback_data=f"att_para_{i}_{subject}"))
    bot.send_message(message.chat.id,
        f"📊 <b>Барлау — {today}</b>\n\nҚай параны белгилейсиз:", reply_markup=markup)

@bot.message_handler(func=lambda m: m.text == "📅 Барлау тарихы")
@check_access
def attendance_history(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT DISTINCT LEFT(date,7) as ym FROM attendance ORDER BY ym DESC")
        months = [r[0] for r in cursor.fetchall()]
    if not months:
        bot.send_message(message.chat.id, "📭 Барлау жазылмаған.", reply_markup=attendance_submenu())
        return
    markup = types.InlineKeyboardMarkup()
    for ym in months:
        y, mo = ym.split("-")
        markup.add(types.InlineKeyboardButton(
            text=f"📅 {MONTHS_RU.get(int(mo), mo)} {y}",
            callback_data=f"hist_month_{ym}"))
    bot.send_message(message.chat.id, "📅 <b>Барлау тарихы</b>\n\nАйды таңлаңыз:", reply_markup=markup)

# ── ӨШІРІУ — FIX: SQL injection whitelist ─────────────────────
@bot.message_handler(func=lambda m: m.text == "🗑 Өшириу")
@check_access
def delete_management(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "🗑 <b>Өшириу бөлими</b>", reply_markup=delete_submenu())

@bot.message_handler(func=lambda m: m.text == "🗑 Материал өшириу")
@check_access
def delete_material_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,file_type,uploader_username,date FROM materials ORDER BY date DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Материаллар жоқ.", reply_markup=delete_submenu())
        return
    text = "🗑 <b>Материалларды өшириу:</b>\n\n"
    for r in rows:
        text += f"ID:<code>{r[0]}</code> | {r[1]} | {'@'+r[2] if r[2] else '—'} | {r[3]}\n"
    text += "\n\nID ямаса <code>all</code> жазыңыз:"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, lambda m: delete_material(m))

@bot.message_handler(func=lambda m: m.text == "🗑 Фото/Видео өшириу")
@check_access
def delete_gallery_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,file_type,uploader_username,date FROM gallery ORDER BY date DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Галерея бос.", reply_markup=delete_submenu())
        return
    text = "🗑 <b>Фото/Видео өшириу:</b>\n\n"
    for r in rows:
        text += f"ID:<code>{r[0]}</code> | {r[1]} | {'@'+r[2] if r[2] else '—'} | {r[3]}\n"
    text += "\n\nID ямаса <code>all</code> жазыңыз:"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, lambda m: delete_gallery_item(m))

@bot.message_handler(func=lambda m: m.text == "🗑 Жаңалық өшириу")
@check_access
def delete_news_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,author_username,date,content FROM user_news ORDER BY date DESC")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Жаңалықлар жоқ.", reply_markup=delete_submenu())
        return
    text = "🗑 <b>Жаңалықларды өшириу:</b>\n\n"
    for r in rows:
        uname = f"@{r[1]}" if r[1] else "Белгисиз"
        preview = r[3][:40] + "..." if len(r[3]) > 40 else r[3]
        text += f"ID:<code>{r[0]}</code> | {uname}\n📌 {preview}\n{'─'*20}\n"
    text += "\n\nID ямаса <code>all</code> жазыңыз:"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, lambda m: delete_news_item(m))

# FIX: SQL injection-дан қорғалған whitelist арқалы өшириу
def _delete_table_item(message, table):
    """
    SQL injection-дан қорғаған улыума өшириу функциясы.
    table параметри тек ALLOWED_DELETE_TABLES ишинде болыуы керек.
    """
    if not is_admin(message.from_user.id):
        return
    # CRITICAL FIX: Тек рұхсат етилген кестелерге рұхсат
    if table not in ALLOWED_DELETE_TABLES:
        logger.error(f"_delete_table_item: рұхсатсыз кесте аты: {table!r}")
        bot.send_message(message.chat.id, "❌ Қате: рұхсатсыз операция.", reply_markup=delete_submenu())
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "🗑 Өшириу бөлими", reply_markup=delete_submenu())
        return
    try:
        with db_cursor() as (conn, cursor):
            if message.text.strip().lower() == "all":
                # Параметрленген кесте аты (whitelist арқалы тексерилди)
                cursor.execute(f"DELETE FROM {table}")  # nosec — whitelist тексерилди
                d = cursor.rowcount
                conn.commit()
                bot.send_message(message.chat.id, f"✅ {d} жазба өширилди.", reply_markup=delete_submenu())
            else:
                try:
                    rid = int(message.text.strip())
                except ValueError:
                    msg = bot.send_message(message.chat.id,
                        "❌ ID ямаса <code>all</code> жазыңыз:", reply_markup=back_menu())
                    bot.register_next_step_handler(msg, lambda m: _delete_table_item(m, table))
                    return
                cursor.execute(f"SELECT id FROM {table} WHERE id=%s", (rid,))  # nosec — whitelist
                if not cursor.fetchone():
                    bot.send_message(message.chat.id, "⚠️ Табылмады.", reply_markup=delete_submenu())
                    return
                cursor.execute(f"DELETE FROM {table} WHERE id=%s", (rid,))  # nosec — whitelist
                conn.commit()
                bot.send_message(message.chat.id, f"✅ ID:{rid} өширилди.", reply_markup=delete_submenu())
    except Exception as e:
        logger.error(f"_delete_table_item({table}): {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=delete_submenu())

def delete_material(message): _delete_table_item(message, "materials")
def delete_gallery_item(message): _delete_table_item(message, "gallery")
def delete_news_item(message): _delete_table_item(message, "user_news")

# ── СТАТИСТИКА, СТУДЕНТЛЕР, т.б. ─────────────────────────────
@bot.message_handler(func=lambda m: m.text in [
    "👥 Студентлер", "❗ Сабақ болмайды", "📈 Статистика", "📩 Ус/Ша келген"])
@check_access
def admin_panel_actions(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫 Сіз админ емессіз!")
        return
    if message.text == "👥 Студентлер":
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT id,username,last_active,full_name FROM students WHERE started=1 "
                "ORDER BY last_active DESC")
            sr = cursor.fetchall()
            cursor.execute(
                "SELECT id,full_name FROM students WHERE started=0 OR started IS NULL "
                "ORDER BY full_name")
            nsr = cursor.fetchall()
        now_t = now_uz()
        oc = sum(1 for r in sr if _is_online(r[2], now_t))
        text = (f"👥 <b>Студентлер дизімі</b>\n✅ Ботқа кірген: <b>{len(sr)}</b>\n"
                f"🟢 Онлайн: <b>{oc}</b> | 🔴 Офлайн: <b>{len(sr)-oc}</b>\n{'─'*30}\n\n")
        if sr:
            text += "📲 <b>Ботқа кіргендер:</b>\n\n"
            for i, r in enumerate(sr, 1):
                uname = f"@{r[1]}" if r[1] else "—"
                name = r[3] or uname
                text += f"{i}. {get_online_status(r[2])}\n   👤 <b>{name}</b>\n   🔗 {uname}\n\n"
        else:
            text += "📭 Әлі ешкім ботқа кірмеген.\n\n"
        if nsr:
            text += f"{'─'*30}\n⏳ <b>Ботқа кірмегендер ({len(nsr)}):</b>\n"
            for r in nsr:
                text += f"  • {r[1] or '—'} (ID: <code>{r[0]}</code>)\n"
        send_long_message(message.chat.id, text, reply_markup=admin_menu())

    elif message.text == "📈 Статистика":
        with db_cursor() as (_, cursor):
            cursor.execute("SELECT COUNT(*) FROM students"); s = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM schedule"); l = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM user_news"); n = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM materials"); mat = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM gallery"); g = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM suggestions"); sg = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(DISTINCT date) FROM attendance"); ad = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM blocked_users"); bl = cursor.fetchone()[0]
        bot.send_message(message.chat.id,
            f"📈 <b>Статистика:</b>\n\n👥 Студентлер: <b>{s}</b>\n📅 Сабақлар: <b>{l}</b>\n"
            f"📰 Жаңалықлар: <b>{n}</b>\n📚 Материаллар: <b>{mat}</b>\n🎞 Галерея: <b>{g}</b>\n"
            f"💡 Ұсыныслар: <b>{sg}</b>\n📊 Барлау күндері: <b>{ad}</b>\n🔒 Блокланған: <b>{bl}</b>",
            reply_markup=admin_menu())

    elif message.text == "📩 Ус/Ша келген":
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT s.content,s.user_id,s.date,st.username "
                "FROM suggestions s LEFT JOIN students st ON s.user_id=st.id "
                "ORDER BY s.date DESC")
            rows = cursor.fetchall()
        if not rows:
            bot.send_message(message.chat.id, "📭 Жоқ.", reply_markup=admin_menu())
            return
        chunks = []
        cur = f"📩 <b>Ұсыныс/Шағымлар ({len(rows)}):</b>\n\n"
        for r in rows:
            entry = (f"👤 {'@'+r[3] if r[3] else 'Белгісіз'} | <code>{r[1]}</code>\n"
                     f"🕐 {r[2]}\n💬 {r[0]}\n{'─'*25}\n")
            if len(cur) + len(entry) > 3800:
                chunks.append(cur)
                cur = ""
            cur += entry
        if cur: chunks.append(cur)
        for i, chunk in enumerate(chunks):
            bot.send_message(message.chat.id, chunk,
                reply_markup=admin_menu() if i == len(chunks) - 1 else None)

    elif message.text == "❗ Сабақ болмайды":
        send_to_students(text="❗ <b>Назар аударыңыз!</b>\nБүгін сабақ болмайды!")
        bot.send_message(message.chat.id, "✅ Жіберілді!", reply_markup=admin_menu())

# ── БАРЛАУ CALLBACKS ──────────────────────────────────────────
def build_attendance_markup(session):
    idx = session["current_index"]
    if idx >= len(session["students"]): return None, None
    student = session["students"][idx]
    sid, sname = student[0], student[1]
    total = len(session["students"])
    done = len(session["results"])
    markup = types.InlineKeyboardMarkup()
    markup.row(
        types.InlineKeyboardButton("✅ Бар", callback_data=f"att_mark_present_{sid}"),
        types.InlineKeyboardButton("❌ Жоқ", callback_data=f"att_mark_absent_{sid}"))
    markup.add(types.InlineKeyboardButton("🏁 Жуумақлау", callback_data="att_finish"))
    text = (f"📊 <b>Барлау — {session['para']}-пара: {session['subject']}</b>\n"
            f"📅 {session['date']}\n{'─'*30}\n👤 <b>{sname}</b>\n{'─'*30}\n"
            f"<i>{done}/{total} белгіленді</i>")
    return text, markup

@bot.callback_query_handler(func=lambda c: c.data.startswith("att_para_"))
@check_access_cb
def att_select_para(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "🚫 Тек admin-ге!")
        return
    parts = call.data.split("_", 3)
    para = int(parts[2])
    subject = parts[3]
    date_str = now_uz().strftime("%Y-%m-%d")
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT id,full_name FROM students WHERE full_name IS NOT NULL AND full_name!='' "
            "ORDER BY full_name")
        students = [[r[0], r[1]] for r in cursor.fetchall()]
    if not students:
        bot.answer_callback_query(call.id, "Студентлер дізімі бос!")
        bot.edit_message_text("📭 Студентлерде ФИО жоқ.",
            call.message.chat.id, call.message.message_id)
        return
    session = {
        "date": date_str, "para": para, "subject": subject,
        "students": students, "results": {}, "current_index": 0}
    save_attendance_session(call.from_user.id, session)
    text, markup = build_attendance_markup(session)
    bot.edit_message_text(text, call.message.chat.id, call.message.message_id,
        reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("att_mark_"))
@check_access_cb
def att_mark_student(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "🚫 Тек admin-ге!")
        return
    session = load_attendance_session(call.from_user.id)
    if not session:
        bot.answer_callback_query(call.id, "Сессия табылмады, қайта бастаңыз.")
        return
    parts = call.data.split("_")
    status = parts[2]
    try:
        sid = int(parts[3])
    except Exception:
        bot.answer_callback_query(call.id, "Қате, қайта бастаңыз.")
        return
    session["results"][sid] = status
    session["current_index"] += 1
    save_attendance_session(call.from_user.id, session)
    sname = next((s[1] for s in session["students"] if s[0] == sid), "—")
    bot.answer_callback_query(call.id, f"{'✅' if status == 'present' else '❌'} {sname}")
    if session["current_index"] >= len(session["students"]):
        finish_attendance(call.message, call.from_user.id)
    else:
        text, markup = build_attendance_markup(session)
        try:
            bot.edit_message_text(text, call.message.chat.id, call.message.message_id,
                reply_markup=markup, parse_mode="HTML")
        except Exception:
            pass

@bot.callback_query_handler(func=lambda c: c.data == "att_finish")
@check_access_cb
def att_finish_early(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id)
        return
    session = load_attendance_session(call.from_user.id)
    if not session:
        bot.answer_callback_query(call.id, "Сессия табылмады.")
        return
    bot.answer_callback_query(call.id, "Барлау жуумақланды!")
    finish_attendance(call.message, call.from_user.id)

def finish_attendance(message, admin_id):
    session = load_attendance_session(admin_id)
    if not session: return
    delete_attendance_session(admin_id)
    date_str = session["date"]
    para = session["para"]
    subject = session["subject"]
    students = session["students"]
    results = session["results"]

    present_list = []
    absent_list = []

    try:
        with db_cursor() as (conn, cursor):
            for item in students:
                sid, sname = item[0], item[1]
                status = results.get(sid, "absent")
                cursor.execute(
                    "INSERT INTO attendance(date,para,subject,student_id,student_name,status) "
                    "VALUES(%s,%s,%s,%s,%s,%s)",
                    (date_str, para, subject, sid, sname, status))
                if status == "present":
                    present_list.append(sname)
                else:
                    absent_list.append((sid, sname))
            conn.commit()
    except Exception as e:
        logger.error(f"finish_attendance DB: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ Барлауды сақтауда қате: {e}")
        return

    total = len(students)
    # FIX: Тек results-та белгіленген absent-тар туралы ескертіу
    actually_marked = len(results)
    result_text = (
        f"📊 <b>Барлау нәтижесі сақталды!</b>\n"
        f"📅 {date_str} | {para}-пара: <b>{subject}</b>\n{'─'*30}\n"
        f"✅ Бар: <b>{len(present_list)}/{total}</b>\n"
        f"❌ Жоқ: <b>{len(absent_list)}/{total}</b>\n")
    if actually_marked < total:
        result_text += f"⚠️ <i>{total - actually_marked} студент белгіленбеді (absent деп саналды)</i>\n"
    result_text += f"{'─'*30}\n"
    if absent_list:
        result_text += "❌ <b>Жоқтар:</b>\n"
        for _, n in absent_list:
            result_text += f"  • {n}\n"
    else:
        result_text += "🎉 Барлық студентлер бар!\n"
    try:
        bot.edit_message_text(result_text, message.chat.id, message.message_id,
            parse_mode="HTML", reply_markup=None)
    except Exception:
        bot.send_message(message.chat.id, result_text, parse_mode="HTML")

    path = generate_attendance_excel(students, results, date_str, para, subject)
    if path:
        if not send_excel_file(message.chat.id, path,
                caption=f"📊 {date_str} | {para}-пара: {subject}"):
            bot.send_message(message.chat.id, "⚠️ Excel жіберілмеді.")
    else:
        bot.send_message(message.chat.id, "⚠️ Excel жасалмады (openpyxl орнатылған ба?)")

    for sid, sname in absent_list:
        try:
            bot.send_message(sid,
                f"⚠️ <b>Ескертіу!</b>\n\nСіз бүгін <b>{para}-парада</b> (<b>{subject}</b>) болмадыңыз!\n"
                f"📅 {date_str}\n\nСебебіңізді группаға хабарлаңыз.")
        except Exception:
            pass
    bot.send_message(message.chat.id,
        "✅ Барлау сақталды!\n📅 Тарихты <b>Барлау тарихы</b> арқалы ашыңыз.",
        reply_markup=attendance_submenu())

@bot.callback_query_handler(func=lambda c: c.data.startswith("hist_month_"))
@check_access_cb
def hist_select_month(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "🚫 Тек admin-ге!")
        return
    ym = call.data.replace("hist_month_", "")
    y, mo = ym.split("-")
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT DISTINCT date FROM attendance WHERE LEFT(date,7)=%s ORDER BY date DESC", (ym,))
        days = [r[0] for r in cursor.fetchall()]
    if not days:
        bot.answer_callback_query(call.id, "Бұл айда барлау жоқ.")
        return
    markup = types.InlineKeyboardMarkup()
    for d in days:
        markup.add(types.InlineKeyboardButton(
            text=f"📆 {date_to_ru(d)}", callback_data=f"hist_day_{d}"))
    markup.add(types.InlineKeyboardButton("◀️ Назад", callback_data="hist_back_months"))
    bot.edit_message_text(
        f"📅 <b>{MONTHS_RU.get(int(mo), mo)} {y}</b>\n\nКүнді таңлаңыз:",
        call.message.chat.id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda c: c.data == "hist_back_months")
@check_access_cb
def hist_back_to_months(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id)
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT DISTINCT LEFT(date,7) as ym FROM attendance ORDER BY ym DESC")
        months = [r[0] for r in cursor.fetchall()]
    markup = types.InlineKeyboardMarkup()
    for ym in months:
        y, mo = ym.split("-")
        markup.add(types.InlineKeyboardButton(
            text=f"📅 {MONTHS_RU.get(int(mo), mo)} {y}",
            callback_data=f"hist_month_{ym}"))
    bot.edit_message_text(
        "📅 <b>Барлау тарихы</b>\n\nАйды таңлаңыз:",
        call.message.chat.id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("hist_day_"))
@check_access_cb
def hist_select_day(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "🚫 Тек admin-ге!")
        return
    date_str = call.data.replace("hist_day_", "")
    ym = date_str[:7]
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT DISTINCT para,subject FROM attendance WHERE date=%s ORDER BY para", (date_str,))
        paras = cursor.fetchall()
    if not paras:
        bot.answer_callback_query(call.id, "Бұл күнде барлау жоқ.")
        return
    markup = types.InlineKeyboardMarkup()
    for para, subject in paras:
        markup.add(types.InlineKeyboardButton(
            text=f"📖 {para}-пара: {subject}",
            callback_data=f"hist_para_{date_str}_{para}"))
    markup.add(types.InlineKeyboardButton("◀️ Назад", callback_data=f"hist_month_{ym}"))
    bot.edit_message_text(
        f"📆 <b>{date_to_ru(date_str)}</b>\n\nПараны таңлаңыз:",
        call.message.chat.id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("hist_para_"))
@check_access_cb
def hist_show_para(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "🚫 Тек admin-ге!")
        return
    # FIX: "hist_para_2024-01-15_2" — соңғы '_' арқылы бөлу
    # split("_") қате береді: ['hist','para','2024-01','15','2']
    rest = call.data[len("hist_para_"):]   # "2024-01-15_2"
    last_us = rest.rfind("_")
    date_str = rest[:last_us]              # "2024-01-15"
    para = int(rest[last_us + 1:])         # 2
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT student_name,status FROM attendance WHERE date=%s AND para=%s ORDER BY student_name",
            (date_str, para))
        records = cursor.fetchall()
        cursor.execute(
            "SELECT DISTINCT subject FROM attendance WHERE date=%s AND para=%s", (date_str, para))
        sr = cursor.fetchone()
    subject = sr[0] if sr else "—"
    present = [r[0] for r in records if r[1] == "present"]
    absent = [r[0] for r in records if r[1] == "absent"]
    total = len(records)
    text = (f"📊 <b>Барлау нәтижесі</b>\n"
            f"📆 {date_to_ru(date_str)} | {para}-пара: <b>{subject}</b>\n{'─'*30}\n"
            f"✅ Бар: <b>{len(present)}/{total}</b>\n❌ Жоқ: <b>{len(absent)}/{total}</b>\n{'─'*30}\n")
    if present:
        text += "✅ <b>Барлар:</b>\n" + "".join(f"  • {n}\n" for n in present) + "\n"
    if absent:
        text += "❌ <b>Жоқтар:</b>\n" + "".join(f"  • {n}\n" for n in absent)
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton(
        "📥 Excel жүклеу", callback_data=f"hist_excel_{date_str}_{para}"))
    markup.add(types.InlineKeyboardButton("◀️ Назад", callback_data=f"hist_day_{date_str}"))
    bot.edit_message_text(text, call.message.chat.id, call.message.message_id,
        reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("hist_excel_"))
@check_access_cb
def hist_download_excel(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "🚫 Тек admin-ге!")
        return
    # FIX: "hist_excel_2024-01-15_2" — соңғы '_' арқылы бөлу
    rest = call.data[len("hist_excel_"):]   # "2024-01-15_2"
    last_us = rest.rfind("_")
    date_str = rest[:last_us]               # "2024-01-15"
    para = int(rest[last_us + 1:])          # 2
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT student_name,status FROM attendance WHERE date=%s AND para=%s ORDER BY student_name",
            (date_str, para))
        records = cursor.fetchall()
        cursor.execute(
            "SELECT DISTINCT subject FROM attendance WHERE date=%s AND para=%s", (date_str, para))
        sr = cursor.fetchone()
    subject = sr[0] if sr else "—"
    path = generate_attendance_excel(records, None, date_str, para, subject)
    if path:
        ok = send_excel_file(call.message.chat.id, path,
            caption=f"📊 Барлау: {date_str} | {para}-пара: {subject}")
        bot.answer_callback_query(call.id, "✅ Excel жіберілді!" if ok else "❌ Жіберу қатесі.")
    else:
        bot.answer_callback_query(call.id, "❌ Excel жасалмады.")

# ── ПӘНЛЕР ───────────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📖 Пәнлер")
@check_access
def show_variants_menu(message):
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT subject,COUNT(*) as cnt FROM test_variants GROUP BY subject ORDER BY subject")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Әлі вариант жүклемеген.",
            reply_markup=main_menu(message.from_user.id))
        return
    markup = types.InlineKeyboardMarkup()
    for subj, cnt in rows:
        markup.add(types.InlineKeyboardButton(
            text=f"📖 {subj} ({cnt})", callback_data=f"var_subj_{subj}"))
    bot.send_message(message.chat.id, "📖 <b>Пәндер</b>\n\nПәнді таңлаңыз:", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data.startswith("var_subj_"))
@check_access_cb
def show_variants_by_subject(call):
    subj = call.data.replace("var_subj_", "")
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT id,file_name,file_type,date FROM test_variants WHERE subject=%s ORDER BY date DESC",
            (subj,))
        rows = cursor.fetchall()
    if not rows:
        bot.answer_callback_query(call.id, "Бұл пәнде файл жоқ.")
        return
    markup = types.InlineKeyboardMarkup()
    for r in rows:
        icon = {"photo": "🖼", "document": "📄", "video": "🎬"}.get(r[2], "📎")
        name = r[1] or f"Файл #{r[0]}"
        markup.add(types.InlineKeyboardButton(
            text=f"{icon} {name}", callback_data=f"var_file_{r[0]}"))
    markup.add(types.InlineKeyboardButton("◀️ Артқа", callback_data="var_back"))
    bot.edit_message_text(f"📖 <b>{subj}</b>\n\nФайлды таңлаңыз:",
        call.message.chat.id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda c: c.data.startswith("var_file_"))
@check_access_cb
def send_variant_file(call):
    try:
        vid = int(call.data.replace("var_file_", ""))
    except ValueError:
        bot.answer_callback_query(call.id, "Қате ID.")
        return
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT file_id,file_type,file_name,subject FROM test_variants WHERE id=%s", (vid,))
        row = cursor.fetchone()
    if not row:
        bot.answer_callback_query(call.id, "Файл табылмады.")
        return
    file_id, file_type, file_name, subject = row
    cap = f"📖 <b>{subject}</b>\n📎 {file_name or ''}"
    try:
        if file_type == "photo":
            bot.send_photo(call.message.chat.id, file_id, caption=cap)
        elif file_type == "video":
            bot.send_video(call.message.chat.id, file_id, caption=cap)
        else:
            bot.send_document(call.message.chat.id, file_id, caption=cap)
        bot.answer_callback_query(call.id, "✅ Жіберілді!")
    except Exception as e:
        bot.answer_callback_query(call.id, f"❌ Қате: {e}")

@bot.callback_query_handler(func=lambda c: c.data == "var_back")
@check_access_cb
def variants_back(call):
    with db_cursor() as (_, cursor):
        cursor.execute(
            "SELECT subject,COUNT(*) as cnt FROM test_variants GROUP BY subject ORDER BY subject")
        rows = cursor.fetchall()
    if not rows:
        bot.edit_message_text("📭 Пәндер жоқ.", call.message.chat.id, call.message.message_id)
        return
    markup = types.InlineKeyboardMarkup()
    for subj, cnt in rows:
        markup.add(types.InlineKeyboardButton(
            text=f"📖 {subj} ({cnt})", callback_data=f"var_subj_{subj}"))
    bot.edit_message_text("📖 <b>Пәндер</b>\n\nПәнді таңлаңыз:",
        call.message.chat.id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
    bot.answer_callback_query(call.id)

# ── ПӘН БАСҚАРЫУ ──────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📖 Пән басқарыу")
@check_access
def panler_admin(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    bot.send_message(message.chat.id, "📖 <b>Пән басқарыу</b>", reply_markup=panler_admin_submenu())

@bot.message_handler(func=lambda m: m.text == "➕ Пән қосыу")
@check_access
def add_variant_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    msg = bot.send_message(message.chat.id, "📖 <b>Пәннің атын жазыңыз:</b>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_variant_subject)

def handle_variant_subject(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📖 Пән басқарыу", reply_markup=panler_admin_submenu())
        return
    subject = message.text.strip()
    if len(subject) < 2 or len(subject) > 100:
        msg = bot.send_message(message.chat.id,
            "❌ Пән атын дұрыс жазыңыз (2-100 таңба):", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_variant_subject)
        return
    set_user_state(message.from_user.id, f"variant:{subject}")
    msg = bot.send_message(message.chat.id,
        f"📤 <b>{subject}</b>\n\nФайл жібериңіз:", reply_markup=back_menu())
    bot.register_next_step_handler(msg, lambda m: handle_variant_file(m, subject))

def handle_variant_file(message, subject):
    if not is_admin(message.from_user.id):
        return
    if message.text and message.text == "⬅️ Артқа":
        clear_user_state(message.from_user.id)
        bot.send_message(message.chat.id, "📖 Пән басқарыу", reply_markup=panler_admin_submenu())
        return
    uid = message.from_user.id
    file_id = None
    file_type = None
    file_name = None
    if message.document:
        file_id = message.document.file_id
        file_type = "document"
        file_name = message.document.file_name or "Файл"
    elif message.photo:
        file_id = message.photo[-1].file_id
        file_type = "photo"
        file_name = "Фото"
    elif message.video:
        file_id = message.video.file_id
        file_type = "video"
        file_name = message.video.file_name or "Видео"
    else:
        msg = bot.send_message(message.chat.id,
            "⚠️ Файл, фото ямаса видео жібериңіз:", reply_markup=back_menu())
        bot.register_next_step_handler(msg, lambda m: handle_variant_file(m, subject))
        return
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO test_variants(subject,file_id,file_type,file_name,uploader_id) "
                "VALUES(%s,%s,%s,%s,%s)",
                (subject, file_id, file_type, file_name, uid))
            conn.commit()
        clear_user_state(uid)
        bot.send_message(message.chat.id,
            f"✅ <b>{subject}</b>\n📎 {file_name} қосылды!", reply_markup=panler_admin_submenu())
    except Exception as e:
        logger.error(f"handle_variant_file: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=panler_admin_submenu())

@bot.message_handler(func=lambda m: m.text == "🗑 Пән өшириу")
@check_access
def delete_variant_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "🚫")
        return
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT id,subject,file_name,file_type FROM test_variants ORDER BY subject")
        rows = cursor.fetchall()
    if not rows:
        bot.send_message(message.chat.id, "📭 Варианттар жоқ.", reply_markup=panler_admin_submenu())
        return
    text = "🗑 <b>Пән/вариант өшириу:</b>\n\n"
    for r in rows:
        icon = {"photo": "🖼", "document": "📄", "video": "🎬"}.get(r[3], "📎")
        text += f"ID:<code>{r[0]}</code> | {r[1]} | {icon} {r[2] or '—'}\n"
    text += "\nID ямаса <code>all</code> жазыңыз:"
    msg = bot.send_message(message.chat.id, text, reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_delete_variant)

def handle_delete_variant(message):
    if not is_admin(message.from_user.id):
        return
    if not message.text or message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📖 Пән басқарыу", reply_markup=panler_admin_submenu())
        return
    try:
        with db_cursor() as (conn, cursor):
            if message.text.strip().lower() == "all":
                cursor.execute("DELETE FROM test_variants")
                d = cursor.rowcount
                conn.commit()
                bot.send_message(message.chat.id,
                    f"✅ {d} вариант өшірілді.", reply_markup=panler_admin_submenu())
            else:
                try:
                    rid = int(message.text.strip())
                except ValueError:
                    msg = bot.send_message(message.chat.id,
                        "❌ ID ямаса <code>all</code> жазыңыз:", reply_markup=back_menu())
                    bot.register_next_step_handler(msg, handle_delete_variant)
                    return
                cursor.execute("SELECT file_name,subject FROM test_variants WHERE id=%s", (rid,))
                row = cursor.fetchone()
                if not row:
                    bot.send_message(message.chat.id, "⚠️ Табылмады.", reply_markup=panler_admin_submenu())
                    return
                cursor.execute("DELETE FROM test_variants WHERE id=%s", (rid,))
                conn.commit()
                bot.send_message(message.chat.id,
                    f"✅ <b>{row[1]} — {row[0]}</b> өшірілді.", reply_markup=panler_admin_submenu())
    except Exception as e:
        logger.error(f"handle_delete_variant: {e}", exc_info=True)
        bot.send_message(message.chat.id, f"❌ DB қатесі: {e}", reply_markup=panler_admin_submenu())

# ── САБАҚ/ЕРТЕҢГЕ ─────────────────────────────────────────────
@bot.message_handler(func=lambda m: m.text == "📊 Сабақ/Ертеңге")
@check_access
def sabak_ertenge(message):
    today = DAYS_EN_TO_RU.get(now_uz().strftime("%A"), "")
    tomorrow_dt = now_uz() + timedelta(days=1)
    tomorrow = DAYS_EN_TO_RU.get(tomorrow_dt.strftime("%A"), "")
    with db_cursor() as (_, cursor):
        cursor.execute("SELECT subject,time FROM schedule WHERE day=%s ORDER BY time", (today,))
        today_lessons = cursor.fetchall()
        cursor.execute("SELECT subject,time FROM schedule WHERE day=%s ORDER BY time", (tomorrow,))
        tomorrow_lessons = cursor.fetchall()
    text = "📊 <b>Сабақ хабары</b>\n\n"
    text += f"📅 <b>Бүгін — {today}:</b>\n"
    if today_lessons:
        for i, r in enumerate(today_lessons, 1):
            text += f"  {i}-пара 🕐 {r[1]} — {r[0]}\n"
    else:
        text += "  📭 Сабақ жоқ\n"
    text += f"\n📅 <b>Ертең — {tomorrow}:</b>\n"
    if tomorrow_lessons:
        for i, r in enumerate(tomorrow_lessons, 1):
            text += f"  {i}-пара 🕐 {r[1]} — {r[0]}\n"
    else:
        text += "  📭 Сабақ жоқ\n"
    bot.send_message(message.chat.id, text, reply_markup=sabak_menu())

@bot.message_handler(func=lambda m: m.text == "✅ Бараман")
@check_access
def sabak_keledi(message):
    bot.send_message(message.chat.id,
        "✅ Жақсы! Жолыңыз болсын! Сабаққа уақытында келіңіз! 💪",
        reply_markup=main_menu(message.from_user.id))

@bot.message_handler(func=lambda m: m.text == "❌ Себеп бар")
@check_access
def sabak_kelmeydi(message):
    uid = message.from_user.id
    set_user_state(uid, "sebep_text")
    msg = bot.send_message(message.chat.id, "❌ <b>Себебіңізді жазыңыз:</b>", reply_markup=back_menu())
    bot.register_next_step_handler(msg, handle_sebep_text)

def handle_sebep_text(message):
    if not user_step_check(message):
        return
    uid = message.from_user.id
    if not message.text or message.text == "⬅️ Артқа":
        clear_user_state(uid)
        bot.send_message(message.chat.id, "📊 Сабақ/Ертеңге", reply_markup=sabak_menu())
        return
    sebep_text = message.text
    set_user_state(uid, f"sebep_file:{sebep_text}")
    msg = bot.send_message(message.chat.id,
        "📎 Файл/фото жібере аласыз (дәлел үшын):", reply_markup=_sebep_file_menu())
    bot.register_next_step_handler(msg, lambda m: handle_sebep_file(m, sebep_text))

def handle_sebep_file(message, sebep_text):
    if not user_step_check(message):
        return
    uid = message.from_user.id
    un = message.from_user.username or f"user{uid}"
    fn = message.from_user.first_name or ""
    ln = message.from_user.last_name or ""
    clear_user_state(uid)
    file_id = None
    file_type = None
    if message.document:
        file_id = message.document.file_id
        file_type = "document"
    elif message.photo:
        file_id = message.photo[-1].file_id
        file_type = "photo"
    elif message.text and message.text == "⏭ Өткізіп жіберіу":
        pass
    elif message.text and message.text == "⬅️ Артқа":
        set_user_state(uid, "sebep_text")
        msg = bot.send_message(message.chat.id,
            "❌ <b>Себебіңізді қайта жазыңыз:</b>", reply_markup=back_menu())
        bot.register_next_step_handler(msg, handle_sebep_text)
        return
    admin_text = (
        f"⚠️ <b>Сабаққа келе алмайтын студент:</b>\n\n"
        f"👤 {fn} {ln}\n🔗 @{un}\n🆔 <code>{uid}</code>\n\n"
        f"📝 <b>Себеп:</b>\n{sebep_text}")
    for aid in ADMIN_IDS:
        try:
            bot.send_message(aid, admin_text)
            if file_id and file_type == "photo":
                bot.send_photo(aid, file_id, caption="📎 Дәлел")
            elif file_id and file_type == "document":
                bot.send_document(aid, file_id, caption="📎 Дәлел")
        except Exception:
            pass
    bot.send_message(message.chat.id,
        "✅ <b>Себебіңіз жіберілді!</b>\nАдминлер хабарланды.",
        reply_markup=main_menu(uid))

# ── AI КӨМЕКШІ ────────────────────────────────────────────────
# FIX: AI тарихы үшін thread-safe queue-based паттерн
_ai_history_lock = Lock()
_ai_chat_history: dict = {}   # {user_id: list of {role, content}}
_ai_last_active: dict = {}    # {user_id: float timestamp}

# FIX: AI_MAX_HISTORY мен AI_CONTEXT_SIZE сәйкесті
AI_MAX_HISTORY = 20      # сақталатын максимум хабарлама саны
AI_CONTEXT_SIZE = 10     # API-ға жіберілетін соңғы хабарлама саны

AI_SYSTEM_PROMPT = (
    "Сен S6-DI-23 группасының ақылды көмекшісіңсен. "
    "Сұрақтарға қысқа, толық және достық түрде жауап бер. "
    "Пайдаланушы қай тілде жазса, сол тілде жауап бер "
    "(қарақалпақша, қазақша, орысша, ағылшынша — бәрі болады). "
    "Егер сұрақ оқуға, сабаққа, университетке байланысты болса — ықыласты жауап бер."
)

def _md_to_html(text: str) -> str:
    text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text, flags=re.DOTALL)
    text = re.sub(r'__(.+?)__', r'<u>\1</u>', text, flags=re.DOTALL)
    text = re.sub(r'(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)', r'<i>\1</i>', text)
    text = re.sub(r'`([^`]+)`', r'<code>\1</code>', text)
    return text

def _ai_try_groq(messages: list) -> str:
    import requests
    api_key = os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        raise ValueError("GROQ_API_KEY жоқ")
    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        json={"model": "llama-3.3-70b-versatile", "messages": messages,
              "max_tokens": 1000, "temperature": 0.7},
        timeout=30)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()

def _ai_try_openai(messages: list) -> str:
    import requests
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        raise ValueError("OPENAI_API_KEY жоқ")
    resp = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        json={"model": "gpt-4o-mini", "messages": messages,
              "max_tokens": 1000, "temperature": 0.7},
        timeout=30)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()

def _ai_try_gemini(user_message: str, history: list) -> str:
    import requests
    api_key = os.environ.get("GOOGLE_API_KEY", "")
    if not api_key:
        raise ValueError("GOOGLE_API_KEY жоқ")
    contents = []
    for msg in history:
        role = "user" if msg["role"] == "user" else "model"
        contents.append({"role": role, "parts": [{"text": msg["content"]}]})
    contents.append({"role": "user", "parts": [{"text": user_message}]})
    resp = requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"gemini-2.0-flash:generateContent?key={api_key}",
        headers={"Content-Type": "application/json"},
        json={
            "system_instruction": {"parts": [{"text": AI_SYSTEM_PROMPT}]},
            "contents": contents,
            "generationConfig": {"maxOutputTokens": 1000, "temperature": 0.7}
        },
        timeout=30)
    resp.raise_for_status()
    return resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()

def ai_ask(user_id: int, user_message: str) -> str:
    # FIX: Race condition жою — тарихты lock ішінде алу және жаңарту
    with _ai_history_lock:
        if user_id not in _ai_chat_history:
            _ai_chat_history[user_id] = []
        # API-ға жіберілетін контекст
        history_snapshot = list(_ai_chat_history[user_id][-AI_CONTEXT_SIZE:])
        _ai_last_active[user_id] = time.time()

    messages = [{"role": "system", "content": AI_SYSTEM_PROMPT}]
    messages.extend(history_snapshot)
    messages.append({"role": "user", "content": user_message})

    answer = None
    for fn, args in [
        (_ai_try_groq, (messages,)),
        (_ai_try_openai, (messages,)),
        (_ai_try_gemini, (user_message, history_snapshot)),
    ]:
        try:
            answer = fn(*args)
            break
        except Exception as e:
            logger.error(f"❌ {fn.__name__} қате: {type(e).__name__}: {e}")

    if not answer:
        return (
            "❌ <b>AI уақытша жұмыс істемейді.</b>\n\n"
            "Барлық 3 сервис (Groq, OpenAI, Gemini) жауап бермеді.\n"
            "Кейінірек қайталаңыз ямаса admin-ге хабарласыңыз."
        )

    # FIX: Жауап алынғаннан кейін lock ішінде тарихты жаңарту
    with _ai_history_lock:
        _ai_chat_history[user_id].append({"role": "user", "content": user_message})
        _ai_chat_history[user_id].append({"role": "assistant", "content": answer})
        # AI_MAX_HISTORY шегінен аспау
        if len(_ai_chat_history[user_id]) > AI_MAX_HISTORY:
            _ai_chat_history[user_id] = _ai_chat_history[user_id][-AI_MAX_HISTORY:]

    return answer

def ai_clear_history_mem(user_id: int):
    with _ai_history_lock:
        _ai_chat_history.pop(user_id, None)
        _ai_last_active.pop(user_id, None)

def cleanup_ai_history():
    now_t = time.time()
    with _ai_history_lock:
        # 2 сағат белсенсіз пайдаланушыларды тазалау
        inactive = [uid for uid, t in _ai_last_active.items() if now_t - t > 7200]
        for uid in inactive:
            _ai_chat_history.pop(uid, None)
            _ai_last_active.pop(uid, None)
    if inactive:
        logger.info(f"AI history cleanup: {len(inactive)} пайдаланушы тазаланды")

@bot.message_handler(func=lambda m: m.text == "🤖 AI Көмекші")
@check_access
def ai_menu(message):
    uid = message.from_user.id
    set_user_state(uid, "ai_chat")
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row("🗑 Тарихты тазалау")
    markup.row("⬅️ Артқа")
    bot.send_message(message.chat.id,
        "🤖 <b>AI Көмекші іске қосылды!</b>\n\n"
        "✏️ Кез-келген сұрағыңызды жазыңыз.\n"
        "🌐 Қай тілде жазсаңыз, сол тілде жауап береді.\n\n"
        "🗑 Тарихты тазалау — жаңа сөйлесу бастау үшын\n"
        "⚡ <i>Groq → OpenAI → Gemini (автоматты резерв)</i>",
        reply_markup=markup)

@bot.message_handler(func=lambda m: m.text == "🗑 Тарихты тазалау"
                     and get_user_state(m.from_user.id) == "ai_chat")
@check_access
def ai_clear_cmd(message):
    ai_clear_history_mem(message.from_user.id)
    bot.send_message(message.chat.id,
        "✅ <b>AI тарихы тазаланды!</b>\nЖаңа сөйлесу басталды.")

@bot.message_handler(
    content_types=["text"],
    func=lambda m: get_user_state(m.from_user.id) == "ai_chat"
                   and m.text not in ("⬅️ Артқа", "🗑 Тарихты тазалау"))
@check_access
def ai_chat_handler(message):
    text = message.text.strip()
    if not text:
        bot.send_message(message.chat.id, "✏️ Сұрағыңызды жазыңыз.")
        return
    bot.send_chat_action(message.chat.id, "typing")
    wait_msg = bot.send_message(message.chat.id, "⏳ <i>AI ойланып жатыр...</i>")
    answer = ai_ask(message.from_user.id, text)
    try:
        bot.delete_message(message.chat.id, wait_msg.message_id)
    except Exception:
        pass
    try:
        bot.send_message(message.chat.id, f"🤖 {_md_to_html(answer)}", parse_mode="HTML")
    except Exception:
        bot.send_message(message.chat.id, f"🤖 {answer}")

# ── SCHEDULER ─────────────────────────────────────────────────
def _reminder_already_sent(key: str) -> bool:
    try:
        with db_cursor() as (_, cursor):
            cursor.execute("SELECT key FROM sent_reminders WHERE key=%s", (key,))
            return cursor.fetchone() is not None
    except Exception:
        return False

def _mark_reminder_sent(key: str):
    try:
        with db_cursor() as (conn, cursor):
            cursor.execute(
                "INSERT INTO sent_reminders(key, sent_at) VALUES(%s,%s) ON CONFLICT(key) DO NOTHING",
                (key, now_uz()))
            conn.commit()
    except Exception as e:
        logger.warning(f"_mark_reminder_sent({key}): {e}")

def auto_scheduler():
    _last_ping = 0
    _last_minute = -1  # Бір минутта тек бір рет орындалу үшын

    while True:
        try:
            now = now_uz()
            h, m_ = now.hour, now.minute
            current_minute = h * 60 + m_

            # ── DB ping: әр 4 минут сайын ──
            ping_now = time.time()
            if ping_now - _last_ping >= 240:
                try:
                    with db_cursor() as (_, cursor):
                        cursor.execute("SELECT 1")
                    _last_ping = ping_now
                except Exception as e:
                    logger.warning(f"DB ping қате: {e}")

            # FIX: Бір минутта тек бір рет орындалу
            if current_minute == _last_minute:
                time.sleep(15)
                continue
            _last_minute = current_minute

            # ── Таңғы хабарлама 07:30 ──
            if h == 7 and m_ == 30:
                key = f"morning_{now.strftime('%Y-%m-%d')}"
                if not _reminder_already_sent(key):
                    try:
                        today = DAYS_EN_TO_RU.get(now.strftime("%A"), "")
                        with db_cursor() as (_, cursor):
                            cursor.execute(
                                "SELECT subject,time FROM schedule WHERE day=%s ORDER BY time",
                                (today,))
                            lessons = cursor.fetchall()
                        msg_ = f"☀️ <b>Қайырлы таң!</b>\n📅 Бүгін: <b>{today}</b>\n\n"
                        if lessons:
                            msg_ += "📖 <b>Бүгінгі сабақлар:</b>\n"
                            for i, r in enumerate(lessons, 1):
                                msg_ += f"  {i}-пара 🕐 {r[1]} — {r[0]}\n"
                        else:
                            msg_ += "📭 Бүгін сабақ жоқ. Демалыңыз! 🎉"
                        send_to_students(text=msg_)
                        _mark_reminder_sent(key)
                        logger.info(f"✅ Таңғы хабарлама жіберілді: {key}")
                    except Exception as e:
                        logger.error(f"Таңғы хабарлама қате: {e}", exc_info=True)

            # ── Туылған күн тексеру 09:00 ──
            elif h == 9 and m_ == 0:
                today_str = now.strftime("%m-%d")
                try:
                    with db_cursor() as (_, cursor):
                        cursor.execute(
                            "SELECT id,full_name,birth_date FROM students "
                            "WHERE birth_date IS NOT NULL AND birth_date!=''")
                        students_ = cursor.fetchall()
                    for sid, sname, bd in students_:
                        try:
                            if not bd: continue
                            bd_str = str(bd).strip()[:10]
                            if not bd_str: continue
                            bd_dt = datetime.strptime(bd_str, "%Y-%m-%d")
                            if bd_dt.strftime("%m-%d") != today_str: continue
                            bday_key = f"birthday_{sid}_{now.strftime('%Y-%m-%d')}"
                            if _reminder_already_sent(bday_key): continue
                            age = now.year - bd_dt.year
                            send_to_students(
                                text=(
                                    f"🎂 <b>Бүгін {sname}-ның туылған күні!</b>\n"
                                    f"🎉 Оған {age} жас толды!\n\n"
                                    "Барлық группа атынан құттықтаймыз! 🎊"))
                            try:
                                bot.send_message(
                                    sid,
                                    f"🎂 <b>Туылған күніңіз мүбәрак болсын!</b>\n"
                                    f"🎉 Сізге {age} жас толды!\n\n"
                                    "S6-DI-23 группасы атынан ең жылы тілектерімізді жолдаймыз! 🎊")
                            except Exception:
                                pass
                            _mark_reminder_sent(bday_key)
                            logger.info(f"🎂 Туылған күн хабарламасы: {sname}")
                        except Exception as e:
                            logger.warning(f"Birthday check ({sname}): {e}")
                except Exception as e:
                    logger.error(f"Birthday scheduler: {e}", exc_info=True)

            # ── Тазалау 03:00 ──
            elif h == 3 and m_ == 0:
                clean_key = f"cleanup_{now.strftime('%Y-%m-%d')}"
                if not _reminder_already_sent(clean_key):
                    try:
                        cleanup_old_sessions()
                        clean_rate_limit()
                        cleanup_ai_history()
                        # AI тарихын DB-де тазалау
                        with db_cursor() as (conn, cursor):
                            cursor.execute(
                                "DELETE FROM ai_history WHERE updated_at < %s",
                                (now_uz() - timedelta(days=30),))
                            conn.commit()
                        # Ескі sent_reminders жазбаларын тазалау
                        with db_cursor() as (conn, cursor):
                            cursor.execute(
                                "DELETE FROM sent_reminders WHERE sent_at < %s",
                                (now_uz() - timedelta(days=7),))
                            deleted = cursor.rowcount
                            conn.commit()
                            if deleted:
                                logger.info(f"sent_reminders: {deleted} ескі жазба өшірілді")
                        _mark_reminder_sent(clean_key)
                        logger.info("✅ Түнгі тазалау аяқталды")
                    except Exception as e:
                        logger.error(f"Тазалау қате: {e}", exc_info=True)

        except Exception as e:
            logger.error(f"auto_scheduler қате: {e}", exc_info=True)

        time.sleep(15)

# ── MAIN ──────────────────────────────────────────────────────
if __name__ == "__main__":
    Thread(target=auto_scheduler, daemon=True).start()

    def run_flask():
        port = int(os.environ.get("PORT", 8080))
        app.run(host="0.0.0.0", port=port, debug=False)

    Thread(target=run_flask, daemon=True).start()
    logger.info("🤖 S6-DI-23 Bot іске қосылды!")

    while True:
        try:
            bot.infinity_polling(skip_pending=True, timeout=60, long_polling_timeout=30)
        except Exception as e:
            logger.error(f"Polling қате: {e}", exc_info=True)
            time.sleep(5)